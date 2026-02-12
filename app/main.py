from fastapi import (
    FastAPI, Request, Form, Depends, Response,
    Cookie, HTTPException, UploadFile, File,
    APIRouter, BackgroundTasks
)
from fastapi.staticfiles import StaticFiles
import re
from fastapi.templating import Jinja2Templates
from fastapi.routing import APIRoute
from fastapi.exception_handlers import http_exception_handler
from fastapi.responses import HTMLResponse, StreamingResponse, PlainTextResponse, JSONResponse
from sqlmodel import SQLModel, Field, Session, create_engine, select, Relationship
from pydantic import BaseModel
from typing import Optional, List, Any, Dict
from fastapi import Query
from datetime import datetime, timedelta, date, timezone
from starlette.responses import RedirectResponse
from starlette.middleware.sessions import SessionMiddleware
from random import randint
from sqlalchemy import func, Column, LargeBinary, Integer, text, and_, or_
from sqlalchemy import text as sa_text
import sqlalchemy as sa
from sqlalchemy.dialects.postgresql import JSONB
from itsdangerous import (
    TimestampSigner, BadSignature, SignatureExpired,
    URLSafeSerializer, URLSafeTimedSerializer
)
import os, smtplib, ssl, socket, traceback, time, sys, base64, hashlib
from Crypto.Cipher import AES
from email.message import EmailMessage
import zipfile
import csv
import itsdangerous
import urllib.parse
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import secrets
import json
from pathlib import Path
from zoneinfo import ZoneInfo
import logging, pathlib
from app.vendors.datahub_client import DatahubClient, DatahubError, encrypt_field, _crypto_selftest, pick_latest_general


# ★ 로그 설정(강제적용)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    force=True,
)

APP_ENV = (os.getenv("APP_ENV", "dev") or "").lower()
if APP_ENV == "prod":
    # uvicorn access 로그(요청 라인) 소음 감소
    logging.getLogger("uvicorn.access").setLevel(logging.WARNING)
    # 템플릿/SQL 등 과다 로거도 억제하고 싶으면 여기서 조절
    # logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)

# ★ 현재 실행 파일 경로/버전 태그 찍기
logging.info("[BOOT] main.py loaded from %s", __file__)
logging.info("[BOOT] CWD=%s PYTHONPATH=%s", os.getcwd(), sys.path)
logging.info("[BOOT] BUILD_TS=%s", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))


# --- DataHub Client ---
print("[BOOT] creating DATAHUB client...")
DATAHUB = DatahubClient()
print("[BOOT] DATAHUB ready")


APP_SECRET = os.environ.get("APP_SECRET", "dev-secret")
ADMIN_USER = os.environ.get("ADMIN_USER")
ADMIN_PASS = os.environ.get("ADMIN_PASS")


NHIS_MAX_LIGHT_FETCH = int(os.getenv("NHIS_MAX_LIGHT_FETCH", "1"))     # light는 기본 1회만
NHIS_FETCH_INTERVAL  = float(os.getenv("NHIS_FETCH_INTERVAL", "2.0"))  # 폴링 간격(초)
NHIS_POLL_MAX_SEC    = int(os.getenv("NHIS_POLL_MAX_SEC", "30"))      # 최대 대기(초)

KST = ZoneInfo("Asia/Seoul")

def to_kst(dt: datetime) -> datetime:
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(KST)

KST = timezone(timedelta(hours=9))
def now_kst():
    return datetime.now(tz=KST)

def kst_date_range_to_utc_datetimes(d_from: date | None, d_to: date | None):
    """
    KST 날짜 구간 [d_from, d_to] (둘 다 포함)을 UTC naive datetime 구간
    [start_utc, end_utc) 로 변환한다.
    DB의 naive UTC(datetime.utcnow())와 비교하기 위해 tzinfo 제거.
    """
    start_kst = datetime(d_from.year, d_from.month, d_from.day, 0, 0, 0, tzinfo=KST) if d_from else None
    end_kst   = datetime(d_to.year,   d_to.month,   d_to.day,   23,59,59,999999, tzinfo=KST) if d_to else None
    if d_to:
        # end exclusive: 다음날 00:00 KST
        end_kst = datetime(d_to.year, d_to.month, d_to.day, 0,0,0, tzinfo=KST) + timedelta(days=1)

    start_utc = start_kst.astimezone(timezone.utc).replace(tzinfo=None) if start_kst else None
    end_utc   = end_kst.astimezone(timezone.utc).replace(tzinfo=None)   if end_kst   else None
    return start_utc, end_utc

def ensure_not_completed(survey_completed: str | None = Cookie(default=None)):
    if survey_completed == "1":
        # 이미 완료된 세션은 설문으로 접근 시 포털로 보냄
        raise HTTPException(status_code=307, detail="completed")

# 질문 로드 (앱 기동시 1회)
QUESTIONS_PATH = Path("app/data/survey_questions.json")
with QUESTIONS_PATH.open("r", encoding="utf-8") as f:
    ALL_QUESTIONS = json.load(f)

# 페이지 그룹: (start_id, end_id)
SURVEY_STEPS = [(1, 8), (9, 15), (16, 23)]


def get_questions_for_step(step: int):
    start_id, end_id = SURVEY_STEPS[step-1]
    return [q for q in ALL_QUESTIONS if start_id <= q["id"] <= end_id]


# ---- Basic app setup ----
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(BASE_DIR)
DB_PATH = os.path.join(ROOT_DIR, "app", "data", "app.db")
os.makedirs(os.path.join(ROOT_DIR, "app", "data"), exist_ok=True)

DATABASE_URL = os.environ.get("DATABASE_URL", f"sqlite:///{DB_PATH}")
engine = create_engine(
    DATABASE_URL, echo=False, pool_pre_ping=True, pool_recycle=300,
    connect_args={
    "keepalives": 1,
    "keepalives_idle": 30,
    "keepalives_interval": 10,
    "keepalives_count": 5,
    },
)


app = FastAPI(title="Diet Survey Prototype")

app.mount("/static", StaticFiles(directory=os.path.join(ROOT_DIR, "app", "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(ROOT_DIR, "app", "templates"))

APP_SECRET = os.environ.get("APP_SECRET", "dev-secret")
signer = TimestampSigner(APP_SECRET)

# 임시 라우트 확인 필요
@app.on_event("startup")
def _on_startup():
    global DATAHUB
    logging.info("[BOOT] startup hook fired")
    try:
        if DATAHUB is None:
            logging.info("[BOOT] startup: creating DATAHUB client again...")
            DATAHUB = DatahubClient()
        # selftest 추가 호출(개발 모드에서만)
        app_env   = (os.getenv("APP_ENV", "dev") or "").strip().lower()
        st_flag   = (os.getenv("DATAHUB_SELFTEST", "1") or "").strip()
        logging.info("[BOOT] startup env: APP_ENV=%s SELFTEST=%s", app_env, st_flag)
        if app_env != "prod" and st_flag == "1":
            _crypto_selftest()
    except Exception as e:
        logging.exception("[BOOT] startup error: %r", e)


@app.exception_handler(HTTPException)
async def completed_redirect_handler(request: Request, exc: HTTPException):
    # 설문 뒤로가기 차단 307 → 메인
    if exc.status_code == 307 and exc.detail == "completed":
        return RedirectResponse(url="/", status_code=307)
       # 관리자 호스트 강제 307 → Location 헤더 그대로 사용
    if exc.status_code == 307 and exc.detail == "admin-host-redirect":
        return RedirectResponse(url=exc.headers.get("Location") or "/", status_code=307)

    # ★ 관리자 보호: 401이면 로그인 화면으로
    if exc.status_code == 401 and (request.url.path or "").startswith("/admin"):
        return RedirectResponse(url="/admin/login", status_code=303)
    return await http_exception_handler(request, exc)

@app.middleware("http")
async def no_store_for_survey(request: Request, call_next):
    p = request.url.path
    if p == "/healthz":
        return await call_next(request)  # 최소 비용 통과
    response = await call_next(request)
    if p.startswith("/survey"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response

# ---- Font registration (optional for Korean) ----
FONT_DIR = os.path.join(ROOT_DIR, "app", "fonts")
if os.path.isdir(FONT_DIR):
    try:
        pdfmetrics.registerFont(TTFont("NotoSansKR", os.path.join(FONT_DIR, "NotoSansKR-Regular.otf")))
        pdfmetrics.registerFont(TTFont("NotoSansKR-Bold", os.path.join(FONT_DIR, "NotoSansKR-Bold.otf")))
        pdfmetrics.registerFontFamily("NotoSansKR", normal="NotoSansKR", bold="NotoSansKR-Bold")
        DEFAULT_FONT = "NotoSansKR"
        DEFAULT_FONT_BOLD = "NotoSansKR-Bold"
    except Exception as e:
        print("폰트 등록 실패:", e)
        DEFAULT_FONT = "Helvetica"
        DEFAULT_FONT_BOLD = "Helvetica-Bold"
else:
    DEFAULT_FONT = "Helvetica"
    DEFAULT_FONT_BOLD = "Helvetica-Bold"

AUTH_COOKIE_NAME = "auth"
AUTH_MAX_AGE = 3600 * 1  # 1 hours

def sign_user(user_id: int) -> str:
    return signer.sign(f"user:{user_id}").decode("utf-8")

def verify_user(token: str) -> int:
    try:
        raw = signer.unsign(token, max_age=AUTH_MAX_AGE).decode("utf-8")
        if not raw.startswith("user:"):
            return -1
        return int(raw.split(":")[1])
    except Exception:
        return -1

# --- 민감값 마스킹 헬퍼 (재사용) --- #
def _mask_phone(s: str) -> str:
    if not s: return ""
    d = re.sub(r"[^0-9]", "", s)
    if len(d) < 7: return "***"
    return d[:3] + "-" + "*"*4 + "-" + d[-4:]

def _mask_birth(s: str) -> str:
    # YYYYMMDD → YYYY-**-**
    if not s: return ""
    d = re.sub(r"[^0-9]", "", s)
    if len(d) >= 8:
        return f"{d[:4]}-**-**"
    return "***"



# ---- Models ----
class User(SQLModel, table=True):
    __tablename__ = "users"
    __table_args__ = {"extend_existing": True}
    id: Optional[int] = Field(default=None, primary_key=True)
    phone_hash: str
    name_enc: Optional[str] = None
    birth_year: Optional[int] = None
    gender: Optional[str] = None
    created_at: datetime = Field(default_factory=now_kst)
    birth_date: Optional[date] = None
    height_cm: Optional[float] = None
    weight_kg: Optional[float] = None

class SurveyResponse(SQLModel, table=True):
    # ▶ 테이블명 고정 + 중복 정의 방지 보호막
    __tablename__ = "surveyresponse"
    __table_args__ = {"extend_existing": True}

    # ▶ PK 반드시 필요
    id: Optional[int] = Field(default=None, primary_key=True)

    # ▶ 기존 필드들 (당신 코드 기준으로 이름 유지)
    respondent_id: Optional[int] = None
    answers_json: Optional[str] = None
    score: Optional[int] = None
    submitted_at: Optional[datetime] = None

    # ▶ NHIS 컬럼: JSONB로 정확히 선언 (dict를 그대로 넣어도 저장됨)
    nhis_json: Optional[Dict[str, Any]] = Field(default=None, sa_column=Column(JSONB))
    nhis_raw:  Optional[Dict[str, Any]] = Field(default=None, sa_column=Column(JSONB))

class Respondent(SQLModel, table=True):
    __tablename__ = "respondent"
    # (선택 안전장치) 이미 같은 테이블이 메타데이터에 있을 경우 재정의 허용
    __table_args__ = {"extend_existing": True}

    id: int | None = Field(default=None, primary_key=True)
    user_id: int = Field(index=True)
    campaign_id: str = Field(default="default")
    status: str = Field(default="draft")
    created_at: datetime = Field(default_factory=now_kst)

    # 인적정보 스냅샷
    applicant_name: str | None = None
    birth_date: date | None = None
    gender: str | None = None
    height_cm: float | None = None
    weight_kg: float | None = None

    # 신청번호(고정 순번) — DB 시퀀스에서 자동 발급
    serial_no: int | None = Field(
        default=None,
        sa_column=Column(
            Integer,
            server_default=text("nextval('respondent_serial_no_seq')"),
            unique=True,
            index=True,
        ),
    )

    client_phone: str | None = None
    partner_id: int | None = None
    is_mapped: bool = Field(default=False)
    # updated_at은 NOT NULL 컬럼이므로 기본값을 now_kst로 강제
    updated_at: datetime = Field(default_factory=now_kst)
    
    #동의서 관련 필드
    agreement_all: bool = Field(default=False)
    agreement_at: datetime | None = None
    report_sent_at: datetime | None = None
    sv_memo: str | None = None
    sv_memo_at: datetime | None = None


class ReportFile(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    survey_response_id: int = Field(index=True)
    filename: str
    content: bytes = Field(sa_column=Column(LargeBinary))
    uploaded_at: datetime = Field(default_factory=now_kst)


class UserAdmin(SQLModel, table=True):
    __tablename__ = "user_admin"
    __table_args__ = {"extend_existing": True}

    id: Optional[int] = Field(default=None, primary_key=True)

    division: Optional[str] = None          # varchar(120)
    department: Optional[str] = None        # varchar(120)
    co_num: Optional[str] = None            # varchar(120) - 사번
    name: Optional[str] = None              # varchar(120)

    phone: str                              # varchar(20), NOT NULL, UNIQUE
    mail: Optional[str] = None              # varchar(200)

    is_active: bool = True                  # boolean, default true

    # ✅ 추가
    supervisor: bool = Field(default=False) # boolean, default false

    created_at: Optional[datetime] = None   # DB에서 now()로 채움
    updated_at: Optional[datetime] = None   # DB에서 now()로 채움

    # 새로 추가한 비밀번호 해시 컬럼 (DB에 password_p 로 생성해 둔 상태)
    password_p: Optional[str] = None


#-- 업체담당자, 고객 매핑 테이블 --#
class PartnerClientMapping(SQLModel, table=True):
    __tablename__ = "partner_client_mapping"
    __table_args__ = {"extend_existing": True}

    id: Optional[int] = Field(default=None, primary_key=True)

    created_at: datetime = Field(default_factory=now_kst)

    partner_id: int                              # user_admin.id
    partner_name: Optional[str] = None
    partner_phone: Optional[str] = None

    client_name: Optional[str] = None
    client_phone: Optional[str] = None

    is_mapped: bool = Field(default=False)
    client_submitted_at: Optional[datetime] = None

class Otp(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True)
    phone: str = Field(index=True)
    code: str
    expires_at: datetime
    consumed: bool = Field(default=False)

def init_db():
    SQLModel.metadata.create_all(engine)

@app.on_event("startup")
def on_startup():
    init_db()

# partner, signup_partner 연계
class Partner(SQLModel, table=True):
    __tablename__ = "partner"
    __table_args__ = {"extend_existing": True}

    id: Optional[int] = Field(default=None, primary_key=True)
    created_at: datetime = Field(default_factory=now_kst)

    p_name: str
    p_slug: Optional[str] = None  # ✅ 추가: URL용 slug (예: globalfm, koreafm)
    is_active: bool = True

    p_num: Optional[str] = None
    p_mail: Optional[str] = None
    p_admin: Optional[str] = None
    p_admin_num: Optional[str] = None


# --- NHIS 저장 헬퍼 (rtoken 없이 rid가 확실할 때) ---
def _save_nhis_to_db_with_id(session, respondent_id: int, picked: dict, raw: dict):
    try:
        stmt = sa_text("""
            UPDATE surveyresponse
               SET nhis_json = :js,
                   nhis_raw  = :raw
             WHERE respondent_id = :rid
        """).bindparams(
            rid=respondent_id,
            js=json.dumps(picked or {}, ensure_ascii=False),
            raw=json.dumps(raw or {}, ensure_ascii=False),
        )
        session.exec(stmt)
        session.commit()
        logging.info("[NHIS][DB] saved-by-id rid=%s (json=%s, raw=%s)",
                     respondent_id, bool(picked), bool(raw))
    except Exception as e:
        logging.error("[NHIS][DB][ERR][by-id] %r", e)


# --- NHIS 저장 헬퍼 (rtoken으로 rid 복구 시도) ---
def _save_nhis_to_db(session, request, picked: dict, raw: dict):
    try:
        rid = None
        try:
            tok = (request.query_params.get("rtoken") or request.cookies.get("rtoken") or "")
            rid = verify_token(tok) if tok else -1
            if rid and rid < 0:
                rid = None
        except Exception:
            rid = None

        if not rid:
            logging.info("[NHIS][DB] skip save: no respondent_id")
            return

        _save_nhis_to_db_with_id(session, rid, picked, raw)
    except Exception as e:
        logging.error("[NHIS][DB][ERR] %r", e)



#---- 공단검진 결과 데이터 가공/저장 헬퍼 (legacy)----#

def pick_latest_general_checkup(nhis_data: dict) -> dict | None:
    """
    데이터허브 응답(data.INCOMELIST[])에서 최근 10년 내 가장 최근 1건을 반환.
    일반검진만 포함(암검진 제외) 규칙이 명시돼 있으면 필드로 구분, 없으면 전체에서 최신.
    """
    items = (nhis_data or {}).get("INCOMELIST") or []
    if not items:
        return None

    def parse_date(y, md):
        # GUNYEAR: "2022", GUNDATE: "11/02" 형태 가정
        try:
            y = int(str(y).strip()[:4])
            m, d = md.split("/")
            return datetime(y, int(m), int(d))
        except Exception:
            return None

    ten_years_ago = datetime.now() - timedelta(days=365*10)
    candidates = []
    for it in items:
        dt = parse_date(it.get("GUNYEAR"), it.get("GUNDATE",""))
        if dt and dt >= ten_years_ago:
            # 일반/암 구분이 따로 온다면 여기서 필터(it.get("TYPE") == "GENERAL" 같은 식)
            candidates.append((dt, it))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]




# ---- Simple helpers ----
def hash_phone(phone: str) -> str:
    # naive demo hash (do NOT use in prod). Replace with SHA-256 + salt.
    return "ph_" + str(abs(hash("SALT::" + phone)))

def get_session():
    with Session(engine) as session:
        yield session

# -- admin페이지 partner 비밀번호 검증 헬퍼 -- #
def verify_partner_password(raw_password: str, stored_hash: str | None) -> bool:
    """
    파트너 비밀번호 검증 헬퍼.
    raw_password : 로그인 폼에서 입력한 비밀번호 (평문 숫자 4~6자리)
    stored_hash  : DB(user_admin.password_p)에 저장된 해시 문자열
    """
    if not raw_password or not stored_hash:
        return False

    SALT = "partner_salt_v1"   # ❗ 서비스 운영 시 환경변수로 분리 권장
    hashed = hashlib.sha256((SALT + raw_password).encode("utf-8")).hexdigest()

    return hashed == stored_hash


#-- NHIS 인증간 고객 정보 저장 헬퍼(업체담당자, 고객 매핑 시 활용) --#
def sync_respondent_contact_from_nhis(
    request: Request,
    session: Session,
    respondent: Respondent,
):
    """
    NHIS 간편인증 시작 시 세션에 보관한
    nhis_start_payload(고객 이름/휴대폰)를 Respondent에 반영한다.
    """
    try:
        payload = (request.session or {}).get("nhis_start_payload") or {}
        uname = (payload.get("USERNAME") or "").strip()
        uphone = payload.get("HPNUMBER") or ""
        uphone_digits = re.sub(r"[^0-9]", "", uphone)

        changed = False

        # 이름: 응답자 이름이 비어있으면 NHIS 이름으로 채움
        if uname and not (respondent.applicant_name or "").strip():
            respondent.applicant_name = uname
            changed = True

        # 고객 휴대폰
        if uphone_digits:
            respondent.client_phone = uphone_digits
            changed = True

        if changed:
            respondent.updated_at = now_kst()
            session.add(respondent)
            session.commit()
    except Exception as e:
        logging.warning("[RESP][SYNC-NHIS] %r", e)


#---- 이메일 접수 알림 헬퍼 ----
def mask_second_char(name: str | None) -> str:
    """신청자 이름의 두 번째 글자를 *로 가림 (한글 포함, 1글자면 그대로)"""
    if not name:
        return ""
    s = list(name)
    if len(s) >= 2:
        s[1] = "*"
    return "".join(s)

def send_submission_email(serial_no: int, applicant_name: str, created_at_kst_str: str):
    host = os.getenv("SMTP_HOST")
    port_env = os.getenv("SMTP_PORT", "").strip()
    user = (os.getenv("SMTP_USER") or "").strip()
    password = (os.getenv("SMTP_PASS") or "").strip()
    mail_from = (os.getenv("SMTP_FROM") or "").strip()
    mail_to = (os.getenv("SMTP_TO") or "").strip()
    timeout = int(os.getenv("SMTP_TIMEOUT", "25"))  # 넉넉히 25초

    # 환경 체크
    if not (host and user and password and mail_from and mail_to):
        print("[EMAIL] SMTP env not configured, skip.")
        return

    # 네이버는 보통 '전체 이메일 주소'로 로그인해야 안정적
    login_user = user if "@" in user else f"{user}@naver.com"

    # 메일 만들기
    msg = EmailMessage()
    msg["Subject"] = f"[GaonnSurvey] 새 문진 접수 #{serial_no}"
    msg["From"] = mail_from
    msg["To"] = mail_to
    body = (
        f"새 문진이 접수되었습니다.\n"
        f"- 일련번호: {serial_no}\n"
        f"- 신청자: {applicant_name or '(미입력)'}\n"
        f"- 접수시각(KST): {created_at_kst_str}\n"
        f"\n관리자 페이지에서 확인하세요."
    )
    msg.set_content(body)

    ctx = ssl.create_default_context()

    def try_587():
        print("[EMAIL] connecting 587 STARTTLS...")
        with smtplib.SMTP(host, 587, timeout=timeout) as s:
            s.set_debuglevel(1)  # SMTP 대화 로그 출력
            s.ehlo()
            s.starttls(context=ctx)
            s.ehlo()
            s.login(login_user, password)
            s.send_message(msg)
        print("[EMAIL] sent OK via 587")

    def try_465():
        print("[EMAIL] connecting 465 SSL...")
        with smtplib.SMTP_SSL(host, 465, timeout=timeout, context=ctx) as s:
            s.set_debuglevel(1)
            s.login(login_user, password)
            s.send_message(msg)
        print("[EMAIL] sent OK via 465")

    # IPv6 경로가 느린 환경에서 타임아웃을 줄이기 위해(선택) IPv4 우선 DNS 확인 로그
    try:
        ipv4s = [ai[4][0] for ai in socket.getaddrinfo(host, None, socket.AF_INET)]
        print(f"[EMAIL] DNS A records (IPv4): {ipv4s}")
    except Exception as _e:
        print("[EMAIL] DNS A lookup failed (non-fatal):", repr(_e))

    # 시도: 587 → 실패 시 465 폴백
    try:
        try_587()
        return
    except Exception as e1:
        print("[EMAIL] 587 failed:", repr(e1))
        traceback.print_exc()

    try:
        try_465()
        return
    except Exception as e2:
        print("[EMAIL] 465 failed:", repr(e2))
        traceback.print_exc()

    print("[EMAIL] send failed: both 587 and 465 attempts failed")


def send_report_email(
    to_email: str,
    partner_name: str,
    applicant_name: str,
    partner_requested_at_kst_str: str,
    pdf_filename: str,
    pdf_bytes: bytes,
):
    print("[EMAIL] send_report_email called",
          "to=", to_email,
          "pdf_filename=", pdf_filename,
          "pdf_bytes.len=", (len(pdf_bytes) if pdf_bytes else None))

    host = os.getenv("SMTP_HOST")
    user = (os.getenv("SMTP_USER") or "").strip()
    password = (os.getenv("SMTP_PASS") or "").strip()
    mail_from = (os.getenv("SMTP_FROM") or "").strip()
    timeout = int(os.getenv("SMTP_TIMEOUT", "25"))

    print("[EMAIL] env check HOST=", host,
          "USER=", bool(user),
          "PASS=", bool(password),
          "FROM=", mail_from)

    if not (host and user and password and mail_from and to_email):
        print("[EMAIL] SMTP env not configured or recipient missing, skip.")
        return False  # ✅ 중요

    login_user = user if "@" in user else f"{user}@naver.com"

    masked_applicant = mask_second_char(applicant_name)
    subject = f"[(주)가온앤] {masked_applicant}님의 영양분석 리포트가 도착했습니다."

    body = (
        f"{partner_name}님 안녕하세요,\n"
        f"(주)가온앤 영양분석서비스 담당자입니다.\n\n"
        f"{partner_requested_at_kst_str}에 분석신청하신 고객 {applicant_name}님의 영양분석 리포트를 전달드립니다.\n"
        f"리포트는 고객님 이외 다른 사람에게 전달되지 않도록 주의해주시기 바랍니다.\n\n"
        f"저희 가온앤 서비스를 신청해주셔서 감사합니다.\n"
        f"앞으로도 양질의 서비스를 제공해드리기 위해 최선을 다하겠습니다.\n\n"
        f"즐거운 하루 보내세요!\n\n"
        f"** 보안을 위해 리포트에는 암호가 적용되어있습니다 **\n"
        f"[암호 : 수검자 생년월일 8자리 YYYYMMDD]\n"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = mail_from
    msg["To"] = to_email
    msg.set_content(body)

    fn = pdf_filename or "report.pdf"
    if not fn.lower().endswith(".pdf"):
        fn += ".pdf"
    msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename=fn)

    ctx = ssl.create_default_context()

    try:
        print("[EMAIL] try 587 STARTTLS...")
        with smtplib.SMTP(host, 587, timeout=timeout) as s:
            s.ehlo()
            s.starttls(context=ctx)
            s.ehlo()
            s.login(login_user, password)
            s.send_message(msg)
        print("[EMAIL] try 587 OK")
        return True
    except Exception as e1:
        print("[EMAIL] 587 failed:", repr(e1))
        traceback.print_exc()

    try:
        print("[EMAIL] try 465 SSL...")
        with smtplib.SMTP_SSL(host, 465, timeout=timeout, context=ctx) as s:
            s.login(login_user, password)
            s.send_message(msg)
        print("[EMAIL] try 465 OK")
        return True
    except Exception as e2:
        print("[EMAIL] 465 failed:", repr(e2))
        traceback.print_exc()

    print("[EMAIL] send failed: both 587 and 465 attempts failed")
    return False  # ✅ 중요

#-- 업체담당자, 고객 매핑 헬퍼 1(업체 담당자가 고객 등록 후 문진 작성 시) --#
def try_auto_map_partner_for_respondent(
    session: Session,
    respondent: Respondent,
):
    """
    respondent에 담긴 고객 정보를 기준으로
    최근 1개월 내 등록된 partner_client_mapping 중에서
    아직 매핑되지 않은(is_mapped = false) 레코드를 찾아
    respondent.partner_id 를 채우고, 양쪽 is_mapped를 True로 변경한다.
    """

    if not respondent:
        return

    client_name = (respondent.applicant_name or "").strip()
    client_phone = (respondent.client_phone or "").strip()
    client_phone_digits = re.sub(r"[^0-9]", "", client_phone)

    if not client_name or not client_phone_digits:
        logging.info(
            "[AUTO-MAP] skip: insufficient client info (id=%s name=%s phone=%s)",
            getattr(respondent, "id", None),
            client_name,
            client_phone,
        )
        return

    one_month_ago = datetime.utcnow() - timedelta(days=31)

    # partner_id가 비어 있어도, 매핑 테이블에서 찾을 수 있음
    mapping = session.exec(
        select(PartnerClientMapping)
        .where(
            PartnerClientMapping.client_name == client_name,
            PartnerClientMapping.client_phone == client_phone_digits,
            PartnerClientMapping.is_mapped == False,  # noqa
            PartnerClientMapping.created_at >= one_month_ago,
        )
        .order_by(PartnerClientMapping.created_at.desc())
    ).first()

    logging.info(
        "[AUTO-MAP] try: resp_id=%s partner_id=%s name=%s phone=%s, mapping_found=%s",
        getattr(respondent, "id", None),
        getattr(respondent, "partner_id", None),
        client_name,
        client_phone_digits,
        bool(mapping),
    )

    if not mapping:
        return

    # 매핑 테이블의 partner_id를 respondent에도 반영
    if not respondent.partner_id:
        respondent.partner_id = mapping.partner_id

    respondent.is_mapped = True
    mapping.is_mapped = True
    
    # 고객 문진 제출 시각을 partner_client_mapping의 client_submitted_at로 복사 (있으면)
    try:
        sr = session.exec(
            select(SurveyResponse)
            .where(SurveyResponse.respondent_id == respondent.id)
            .order_by(SurveyResponse.submitted_at.desc())
        ).first()
        if sr and sr.submitted_at:
            mapping.client_submitted_at = sr.submitted_at
    except Exception as e:
        logging.warning(
            "[AUTO-MAP] failed to set client_submitted_at: resp_id=%s err=%r",
            getattr(respondent, "id", None),
            e,
        )
        
    # ✅ 매핑이 실제로 일어난 시점 기록
    respondent.updated_at = now_kst()

    session.add(respondent)
    session.add(mapping)
    session.commit()

    logging.info(
        "[AUTO-MAP] done: resp_id=%s partner_id=%s mapping_id=%s",
        respondent.id,
        respondent.partner_id,
        mapping.id,
    )

# -- 업체 담당자, 고객 매핑 헬퍼 2 (문진 먼저 하고 나중에 담당자가 고객 등록 시) --#
def try_auto_map_respondent_for_mapping(
    session: Session,
    mapping: PartnerClientMapping,
):
    """
    partner_client_mapping 한 건을 기준으로
    최근 1개월 내 생성된 respondent 중에서
    이름/전화가 일치하고 아직 매핑 안 된(is_mapped = false) 건이 있으면
    respondent.partner_id / is_mapped 를 채우고, mapping.is_mapped 도 True 로 변경.
    """
    if not mapping:
        return

    client_name = (mapping.client_name or "").strip()
    client_phone = (mapping.client_phone or "").strip()
    client_phone_digits = re.sub(r"[^0-9]", "", client_phone)

    if not client_name or not client_phone_digits:
        logging.info(
            "[AUTO-MAP2] skip: insufficient mapping client info (mapping_id=%s name=%s phone=%s)",
            getattr(mapping, "id", None),
            client_name,
            client_phone,
        )
        return

    one_month_ago = datetime.utcnow() - timedelta(days=31)

    resp = session.exec(
        select(Respondent)
        .where(
            Respondent.applicant_name == client_name,
            Respondent.client_phone == client_phone_digits,
            Respondent.is_mapped == False,  # noqa
            Respondent.created_at >= one_month_ago,
        )
        .order_by(Respondent.created_at.desc())
    ).first()

    logging.info(
        "[AUTO-MAP2] try: mapping_id=%s partner_id=%s client_name=%s client_phone=%s, resp_found=%s",
        getattr(mapping, "id", None),
        getattr(mapping, "partner_id", None),
        client_name,
        client_phone_digits,
        bool(resp),
    )

    if not resp:
        return

    # respondent 쪽에 partner_id 없으면 채워줌
    if not resp.partner_id:
        resp.partner_id = mapping.partner_id

    resp.is_mapped = True
    mapping.is_mapped = True

    # 고객 문진 제출 시각을 partner_client_mapping의 client_submitted_at로 복사 (있으면)
    try:
        sr = session.exec(
            select(SurveyResponse)
            .where(SurveyResponse.respondent_id == resp.id)
            .order_by(SurveyResponse.submitted_at.desc())
        ).first()
        if sr and sr.submitted_at:
            mapping.client_submitted_at = sr.submitted_at
    except Exception as e:
        logging.warning(
            "[AUTO-MAP2] failed to set client_submitted_at: resp_id=%s err=%r",
            getattr(resp, "id", None),
            e,
        )

    session.add(resp)
    session.add(mapping)
    session.commit()

    logging.info(
        "[AUTO-MAP2] done: resp_id=%s partner_id=%s mapping_id=%s",
        resp.id,
        resp.partner_id,
        mapping.id,
    )


# ---- OTP helpers ----
def issue_otp(session: Session, phone: str) -> str:
    code = f"{randint(0, 999999):06d}"
    otp = Otp(phone=phone, code=code, expires_at=datetime.utcnow() + timedelta(minutes=5))
    session.add(otp)
    session.commit()
    print(f"[OTP] {phone} -> code: {code} (5분 유효)")  # prototype: replace with SMS provider
    return code

def verify_otp(session: Session, phone: str, code: str) -> bool:
    stmt = select(Otp).where(Otp.phone == phone, Otp.code == code, Otp.consumed == False)
    otp = session.exec(stmt).first()
    if not otp:
        return False
    if otp.expires_at < datetime.utcnow():
        return False
    otp.consumed = True
    session.add(otp)
    session.commit()
    return True

# ---- Routes ----
def _host(request: Request) -> str:
    return (request.headers.get("host") or "").split(":")[0].lower()

ADMIN_HOST = "admin.gaonnsurvey.store"
DEFAULT_CAMPAIGN_ID = "demo"

def get_partner_list(session: Session) -> list[str]:
    rows = session.exec(
        sa_text("SELECT p_name FROM partner WHERE is_active = TRUE ORDER BY p_name ASC")
    ).all()
    out: list[str] = []
    for r in rows:
        if not r:
            continue
        nm = (r[0] or "").strip()
        if nm:
            out.append(nm)
    return out


def get_partner_by_slug(session: Session, p_slug: str) -> dict | None:
    row = session.exec(
        sa_text("""
            SELECT p_name, p_slug
              FROM partner
             WHERE p_slug = :s
               AND is_active = TRUE
             LIMIT 1
        """).bindparams(s=p_slug)
    ).first()
    if not row:
        return None
    return {"p_name": (row[0] or "").strip(), "p_slug": (row[1] or "").strip()}


def get_partner_slug_by_name(session: Session, p_name: str) -> str | None:
    row = session.exec(
        sa_text("""
            SELECT p_slug
              FROM partner
             WHERE p_name = :n
             LIMIT 1
        """).bindparams(n=p_name)
    ).first()
    if not row:
        return None
    slug = (row[0] or "").strip()
    return slug or None


@app.get("/info", response_class=HTMLResponse)
def info_form(request: Request, auth: str | None = Cookie(default=None, alias=AUTH_COOKIE_NAME)):
    user_id = verify_user(auth) if auth else -1
    if user_id < 0:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse("info.html", {"request": request})

# -----------------------------------------------
# NHIS 건강검진 조회 페이지 (info 전 단계)
# -----------------------------------------------

@app.get("/nhis", response_class=HTMLResponse)
def nhis_page(
    request: Request,
    auth: str | None = Cookie(default=None, alias=AUTH_COOKIE_NAME),
    session: Session = Depends(get_session),
):
    # 1) 로그인 검사
    user_id = verify_user(auth) if auth else -1
    if user_id < 0:
        # 유효한 로그인 코드 없이 들어오면 /login 으로 보냄
        return RedirectResponse(url="/login", status_code=302)

    # (선택) User가 실제로 존재하는지 한 번 더 확인하고 싶으면:
    user = session.get(User, user_id)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # 2) 정상일 때만 페이지 렌더
    auth_base = os.getenv("DATAHUB_API_BASE", "https://datahub-dev.scraping.co.kr").rstrip("/")
    return templates.TemplateResponse(
        "nhis_fetch.html",
        {
            "request": request,
            "next_url": "/survey",
            "datahub_auth_base": auth_base
        }
    )
    
@app.get("/healthz")
def healthz():
    return PlainTextResponse("ok", status_code=200)

@app.post("/info")
async def info_submit(
    request: Request,
    name: str = Form(...),
    birth_date: str = Form(...),   # "YYMMDD"
    gender: str = Form(...),       # "남" | "여"
    height_cm: str = Form(None),
    weight_kg: str = Form(None),
    auth: str | None = Cookie(default=None, alias=AUTH_COOKIE_NAME),
    session: Session = Depends(get_session),
):
    user_id = verify_user(auth) if auth else -1
    if user_id < 0:
        return RedirectResponse(url="/login", status_code=302)

    # 간단 검증
    try:
        bd = datetime.strptime(birth_date, "%Y-%m-%d").date()
    except:
        return templates.TemplateResponse("error.html", {"request": request, "message": "생년월일 형식이 올바르지 않습니다(YYMMDD)."}, status_code=400)
    if gender not in ("남", "여"):
        return templates.TemplateResponse("error.html", {"request": request, "message": "성별은 남/여 중 선택해주세요."}, status_code=400)

    # 숫자 파싱(선택)
    def to_float(s):
        try:
            return float(s) if s not in (None, "") else None
        except:
            return None
    h_cm = to_float(height_cm)
    w_kg = to_float(weight_kg)

    # 다음 설문 세션 Respondent에 스냅샷으로 옮길 수 있도록 User에도 저장(옵션)
    user = session.get(User, user_id)
    if user:
        user.name_enc = name.strip()
        user.gender = gender.strip()
        user.birth_year = bd.year                       # 호환용 유지
        user.birth_date = bd                            # 실제 생년월일 저장
        user.height_cm = h_cm
        user.weight_kg = w_kg
        session.add(user)
        session.commit()   

    # 설문 시작
    return RedirectResponse(url="/survey", status_code=303)

# --- Session (admin 인증 단일 쿠키) ---
SESSION_MAX_AGE = 60 * 180  # 3시간

app.add_middleware(
    SessionMiddleware,
    secret_key=APP_SECRET,   # (이미 위쪽에 APP_SECRET가 있음)
    max_age=SESSION_MAX_AGE, # 초 단위
    same_site="none",        # 서브도메인/리다이렉트 고려
    https_only=True          # Secure
)

@app.middleware("http")
async def rolling_session_middleware(request: Request, call_next):
    response = await call_next(request)

    now = int(datetime.now(timezone.utc).timestamp())

    # admin rolling
    if request.session.get("admin"):
        issued_at = int(request.session.get("_iat", 0))
        if now - issued_at > (SESSION_MAX_AGE - 300):
            request.session["_iat"] = now  # Set-Cookie 재발급 유도

    # partner rolling
    if request.session.get("partner_id"):
        issued_at_p = int(request.session.get("_iat_partner", 0))
        if now - issued_at_p > (SESSION_MAX_AGE - 300):
            request.session["_iat_partner"] = now

    return response

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    if _host(request) == ADMIN_HOST:
        # 관리자 서브도메인으로 들어오면 관리자 로그인으로 보냄
        return RedirectResponse(url="/admin-portal", status_code=302)
    # 기존 사용자용 홈 유지
    return templates.TemplateResponse("index.html", {"request": request, "apply_url": "/login"})


#portal_home 렌더
@app.get("/admin-portal", response_class=HTMLResponse)
def admin_portal_home(request: Request):
    return templates.TemplateResponse("admin/portal_home.html", {"request": request})

# ---------------------------
# 파트너 로그인 (GET)
# ---------------------------
@app.get("/partner/login", response_class=HTMLResponse)
def partner_login_get(request: Request):
    return templates.TemplateResponse("partner/login.html", {
        "request": request,
        "error": None
    })
    
# -- 파트너 로그아웃 --#
@app.get("/partner/logout")
def partner_logout(request: Request):
    """
    파트너 세션만 정리하고 /partner/login 으로 돌려보낸다.
    (사용자 설문용 AUTH 쿠키는 건드리지 않음)
    """
    request.session.pop("partner_id", None)
    request.session.pop("_iat_partner", None)
    return RedirectResponse(url="/partner/login", status_code=303)



@app.get("/partner/signup", response_class=HTMLResponse)
def partner_signup_form(
    request: Request,
    session: Session = Depends(get_session),
):
    partner_list = get_partner_list(session)
    return templates.TemplateResponse(
        "partner/signup.html",
        {
            "request": request,
            "error": None,
            "message": None,
            "partner_list": partner_list,
        },
    )



# -- 파트너 회원가입 라우트 POST -- #
@app.post("/partner/signup", response_class=HTMLResponse)
async def partner_signup_submit(
    request: Request,
    emp_no: str = Form(...),
    name: str = Form(...),
    phone: str = Form(...),
    email: str = Form(...),
    email_confirm: str = Form(...),
    division: str = Form(""),
    department: str = Form(""),
    password: str = Form(...),
    password_confirm: str = Form(...),
    session: Session = Depends(get_session),
):
    partner_list = get_partner_list(session)

    emp_no = (emp_no or "").strip()
    name = (name or "").strip()
    phone_raw = "".join(c for c in (phone or "") if c.isdigit())
    email = (email or "").strip()
    email_confirm = (email_confirm or "").strip()
    division = (division or "").strip()
    department = (department or "").strip()
    password = (password or "").strip()
    password_confirm = (password_confirm or "").strip()

    #필수값 체크
    error = None
    if not emp_no or not name or not phone_raw or not email or not password or not password_confirm:
        error = "필수 항목을 모두 입력해주세요."
    elif not division:
        error = "소속(회사명 등)을 선택해주세요."
    elif division not in partner_list:
        error = "유효하지 않은 소속(회사명)입니다."
    elif email != email_confirm:
        error = "이메일주소를 확인해주세요"
    elif password != password_confirm:
        error = "비밀번호와 비밀번호 재확인이 일치하지 않습니다."
    elif len(password) < 4 or len(password) > 6 or not password.isdigit():
        error = "비밀번호는 숫자 4~6자리만 가능합니다."
    elif len(phone_raw) < 10 or len(phone_raw) > 11:
        error = "전화번호는 숫자 10~11자리로 입력해주세요."

    #입력값 유지 렌더
    if error:
        return templates.TemplateResponse(
            "partner/signup.html",
            {
                "request": request,
                "error": error,
                "message": None,
                "partner_list": partner_list,
                "emp_no": emp_no,
                "name": name,
                "phone": phone_raw,
                "email": email,
                "email_confirm": email_confirm,
                "division": division,
                "department": department,
            },
        )

    #전화번호 체크
    row = session.exec(
        sa_text("""
            SELECT id
              FROM user_admin
             WHERE phone = :p
             LIMIT 1
        """).bindparams(p=phone_raw)
    ).first()

    if row:
        return templates.TemplateResponse(
            "partner/signup.html",
            {
                "request": request,
                "error": "이미 등록된 전화번호입니다. 로그인을 시도해 주세요.",
                "message": None,
                "partner_list": partner_list,
                "emp_no": emp_no,
                "name": name,
                "phone": phone_raw,
                "email": email,
                "email_confirm": email_confirm,
                "division": division,
                "department": department,
            },
        )

    #비밀번호 처리
    SALT = "partner_salt_v1"
    password_p = hashlib.sha256((SALT + password).encode("utf-8")).hexdigest()

    session.exec(
        sa_text("""
            INSERT INTO user_admin
                (division, department, co_num, name, phone, mail, is_active, password_p)
            VALUES
                (:division, :department, :co_num, :name, :phone, :mail, TRUE, :password_p)
        """).bindparams(
            division=division,
            department=department,
            co_num=emp_no,
            name=name,
            phone=phone_raw,
            mail=email,
            password_p=password_p,
        )
    )
    session.commit()

    #가입 성공, 리다이렉트
    return RedirectResponse(url="/partner/login?msg=signup_ok", status_code=303)

# ---------------------------
# 파트너 로그인 (POST)
# ---------------------------
@app.post("/partner/login", response_class=HTMLResponse)
def partner_login_post(
    request: Request,
    phone: str = Form(...),
    password: str = Form(...),
    session: Session = Depends(get_session)
):
    # 업체 담당자 DB: user_admin
    user = session.exec(
        select(UserAdmin).where(UserAdmin.phone == phone, UserAdmin.is_active == True)
    ).first()

    if not user or not verify_partner_password(password, user.password_p):
        return templates.TemplateResponse("partner/login.html", {
            "request": request,
            "error": "전화번호 또는 비밀번호가 올바르지 않습니다."
        })

    # 세션 값 등록 (관리자 세션과 충돌 방지)
    request.session["partner_id"] = user.id
    request.session["_iat_partner"] = int(datetime.now(timezone.utc).timestamp())

    # 로그인 성공 → 파트너 대시보드로
    return RedirectResponse(url="/partner/dashboard", status_code=302)


#파트너 대시보드 진입점 추가
@app.get("/partner/dashboard", response_class=HTMLResponse)
def partner_dashboard(request: Request):
    if not request.session.get("partner_id"):
        return RedirectResponse(url="/partner/login", status_code=302)

    return templates.TemplateResponse("partner/dashboard.html", {
        "request": request
    })


# -- 파트너 회원정보 수정 라우트 --#
@app.get("/partner/profile", response_class=HTMLResponse)
def partner_profile_get(
    request: Request,
    session: Session = Depends(get_session),
):
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=302)

    user = session.get(UserAdmin, partner_id)
    if not user:
        # 세션은 있는데 DB에 없으면 로그인부터 다시
        request.session.clear()
        return RedirectResponse(url="/partner/login", status_code=302)

    return templates.TemplateResponse(
        "partner/profile.html",
        {
            "request": request,
            "user": user,
            "error": None,
            "message": None,
        },
    )

#-- 파트너 회원정보 수정 POST --#
@app.post("/partner/profile", response_class=HTMLResponse)
async def partner_profile_post(
    request: Request,
    emp_no: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    division: str = Form(""),
    department: str = Form(""),
    current_password: str = Form(""),
    new_password: str = Form(""),
    new_password_confirm: str = Form(""),
    session: Session = Depends(get_session),
):
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=302)

    user = session.get(UserAdmin, partner_id)
    if not user:
        request.session.clear()
        return RedirectResponse(url="/partner/login", status_code=302)

    # 공통 처리: 양쪽 공백/포맷 정리
    emp_no = (emp_no or "").strip()
    phone_raw = "".join(c for c in (phone or "") if c.isdigit())
    email = (email or "").strip()
    division = (division or "").strip()
    department = (department or "").strip()
    current_password = (current_password or "").strip()
    new_password = (new_password or "").strip()
    new_password_confirm = (new_password_confirm or "").strip()

    error = None

    # ---------------------------
    # 1) 비밀번호 변경 의도가 있는지 확인
    # ---------------------------
    wants_pw_change = bool(current_password or new_password or new_password_confirm)

    if wants_pw_change:
        # 1-1) 세 필드 모두 채워져 있어야 함
        if not current_password or not new_password or not new_password_confirm:
            error = "비밀번호 변경 시 현재 비밀번호, 새 비밀번호, 새 비밀번호 확인을 모두 입력해주세요."
        # 1-2) 현재 비밀번호 확인
        elif not verify_partner_password(current_password, user.password_p):
            error = "현재 비밀번호가 올바르지 않습니다."
        # 1-3) 새 비밀번호 규칙 체크
        elif not new_password.isdigit() or not (4 <= len(new_password) <= 6):
            error = "새 비밀번호는 숫자 4~6자리만 가능합니다."
        # 1-4) 새 비밀번호 일치 확인
        elif new_password != new_password_confirm:
            error = "새 비밀번호와 새 비밀번호 확인이 일치하지 않습니다."

    # 전화번호 형식 간단 체크 (필수는 아님)
    if not error and phone_raw and (len(phone_raw) < 10 or len(phone_raw) > 11):
        error = "전화번호는 숫자 10~11자리로 입력해주세요."

    if error:
        return templates.TemplateResponse(
            "partner/profile.html",
            {
                "request": request,
                "user": user,
                "error": error,
                "message": None,
            },
        )

    # ---------------------------
    # 2) 여기까지 왔다면: 비밀번호 조건 OK (또는 변경 안 함)
    #    → 일반 정보 먼저 업데이트
    # ---------------------------
    user.co_num = emp_no or user.co_num
    if phone_raw:
        user.phone = phone_raw
    user.mail = email or user.mail
    user.division = division
    user.department = department

    # ---------------------------
    # 3) 비밀번호 변경 의도가 있고 검증도 통과한 경우에만 password_p 변경
    # ---------------------------
    if wants_pw_change:
        SALT = "partner_salt_v1"
        user.password_p = hashlib.sha256(
            (SALT + new_password).encode("utf-8")
        ).hexdigest()

    session.add(user)
    session.commit()

    # 저장 후 대시보드로 이동
    return RedirectResponse(url="/partner/dashboard", status_code=303)

#-- 고객 담당자와 매핑하기 (고객 서비스 신청하기) --#
@app.get("/partner/mapping", response_class=HTMLResponse)
def partner_mapping_get(
    request: Request,
    session: Session = Depends(get_session),
):
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=302)

    user = session.get(UserAdmin, partner_id)
    if not user:
        request.session.clear()
        return RedirectResponse(url="/partner/login", status_code=302)

    return templates.TemplateResponse(
        "partner/mapping.html",
        {
            "request": request,
            "partner_phone": user.phone,
            "error": None,
            "message": None,
        },
    )

# -- 고객과 담당자 매핑하기 POST --#
@app.post("/partner/mapping", response_class=HTMLResponse)
async def partner_mapping_post(
    request: Request,
    background_tasks: BackgroundTasks,
    client_name: str = Form(...),
    client_phone: str = Form(...),
    session: Session = Depends(get_session),
):

    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=302)

    user = session.get(UserAdmin, partner_id)
    if not user:
        request.session.clear()
        return RedirectResponse(url="/partner/login", status_code=302)

    client_name = (client_name or "").strip()
    client_phone_raw = "".join(c for c in (client_phone or "") if c.isdigit())
    partner_phone = user.phone

    error = None
    message = None

    if not client_name or not client_phone_raw:
        error = "고객 이름과 휴대폰번호를 모두 입력해주세요."
    elif len(client_phone_raw) < 10 or len(client_phone_raw) > 11:
        error = "고객 휴대폰번호는 숫자 10~11자리로 입력해주세요."

    if error:
        return templates.TemplateResponse(
            "partner/mapping.html",
            {
                "request": request,
                "partner_phone": partner_phone,
                "error": error,
                "message": None,
            },
        )

    one_month_ago = datetime.utcnow() - timedelta(days=31)

    # 최근 1개월 내 중복 요청 여부 확인
    dup_row = session.exec(
        sa_text(
            """
            SELECT id
              FROM partner_client_mapping
             WHERE partner_id   = :pid
               AND client_name  = :cname
               AND client_phone = :cphone
               AND created_at  >= :from_dt
             ORDER BY created_at DESC
             LIMIT 1
            """
        ).bindparams(
            pid=partner_id,
            cname=client_name,
            cphone=client_phone_raw,
            from_dt=one_month_ago,
        )
    ).first()

    if dup_row:
        # 이미 등록된 요청이 있으면 그 레코드를 가져와서 재사용
        mapping = session.get(PartnerClientMapping, dup_row[0])
        message = "이미 최근에 등록된 고객 매핑 요청이 있습니다."
    else:
        # 새 매핑 요청 INSERT
        mapping = PartnerClientMapping(
            partner_id=partner_id,
            partner_name=user.name,
            partner_phone=partner_phone,
            client_name=client_name,
            client_phone=client_phone_raw,
            is_mapped=False,
        )
        session.add(mapping)
        session.commit()
        session.refresh(mapping)
        message = "고객 매핑 요청을 등록했습니다."

    # 👉 이 시점에, 이미 존재하는 respondent와 자동 매핑 시도
    prev_is_mapped = bool(getattr(mapping, "is_mapped", False))
    try_auto_map_respondent_for_mapping(session, mapping)

    # 매핑 성공 여부에 따라 메시지 보완 (선택)
    if mapping.is_mapped:
        message = "고객 매핑 요청을 등록했고, 기존 문진과 자동으로 매칭되었습니다."

    # ── 알림메일: '매핑이 새로 완료' 되었고, 고객 문진이 이미 제출된 경우에만 발송 ──
    try:
        if mapping and mapping.is_mapped and (not prev_is_mapped):
            client_name_norm = (mapping.client_name or "").strip()
            client_phone_digits = "".join(c for c in (mapping.client_phone or "") if c.isdigit())

            resp = session.exec(
                select(Respondent)
                .where(
                    Respondent.applicant_name == client_name_norm,
                    Respondent.client_phone == client_phone_digits,
                    Respondent.partner_id == mapping.partner_id,
                )
                .order_by(Respondent.created_at.desc())
            ).first()

            sr = None
            if resp:
                sr = session.exec(
                    select(SurveyResponse)
                    .where(SurveyResponse.respondent_id == resp.id)
                    .order_by(SurveyResponse.submitted_at.desc())
                ).first()

            # 제출(문진 완료) 여부는 SurveyResponse.submitted_at 기준으로 판단
            if resp and sr and sr.submitted_at and (resp.serial_no or 0) > 0:
                created_at_kst_str = (
                    to_kst(resp.created_at).strftime("%Y-%m-%d %H:%M")
                    if resp.created_at else now_kst().strftime("%Y-%m-%d %H:%M")
                )
                background_tasks.add_task(
                    send_submission_email,
                    resp.serial_no or 0,
                    (resp.applicant_name or ""),
                    created_at_kst_str,
                )
    except Exception as e:
        print("[EMAIL][ERR][PARTNER-MAP]", repr(e))


    return templates.TemplateResponse(
        "partner/mapping.html",
        {
            "request": request,
            "partner_phone": partner_phone,
            "error": None,
            "message": message,
        },
    )

# ---------------------------
# 파트너 신청내역 조회 (/partner/requests)
# ---------------------------
@app.get("/partner/requests", response_class=HTMLResponse)
def partner_requests(
    request: Request,
    session: Session = Depends(get_session),
):
    # 1) 로그인 체크
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=302)

    PAGE_SIZE = 50

    # ---------------------------
    # 쿼리스트링에서 필터값 직접 꺼내기
    # ---------------------------
    qp = request.query_params

    raw_page = qp.get("page", "1")
    try:
        page = int(raw_page)
    except ValueError:
        page = 1
    if page < 1:
        page = 1

    date_from = (qp.get("date_from") or "").strip()
    date_to = (qp.get("date_to") or "").strip()
    client_name = (qp.get("client_name") or "").strip()
    client_phone_suffix = (qp.get("client_phone_suffix") or "").strip()
    status = (qp.get("status") or "").strip()
    #문진 미제출
    only_not_submitted = (qp.get("only_not_submitted") or "").strip() in ("1", "true", "on", "yes")


    # 쿼리스트링이 완전히 비어있으면 "첫 진입"으로 간주
    is_first_visit = (request.url.query == "")

    # 오늘 날짜 문자열
    today_str = now_kst().date().isoformat()

    # ---------------------------
    # 날짜 기본값 세팅 로직
    # ---------------------------
    if is_first_visit:
        # 👉 첫 진입: 둘 다 오늘 날짜로 강제 세팅
        date_from_str = today_str
        date_to_str = today_str
    else:
        # 👉 조회 후: 사용자가 넘긴 값을 그대로 유지
        #    (빈 값이면 빈 값 그대로)
        date_from_str = date_from
        date_to_str = date_to

    rows: list[tuple[PartnerClientMapping, Optional[Respondent], Optional[SurveyResponse], Optional[ReportFile]]] = []
    total = 0
    total_pages = 1

    # ---------------------------
    # 여기부터는 date_from_str/date_to_str 기준으로 항상 조회
    # ---------------------------

    def parse_date(s: str | None):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    d_from = parse_date(date_from_str)
    d_to = parse_date(date_to_str)

    # 기간 검사 (둘 다 있는 경우만)
    if d_from and d_to:
        if d_from > d_to:
            return templates.TemplateResponse(
                "partner/requests.html",
                {
                    "request": request,
                    "rows": [],
                    "page": page,
                    "page_size": PAGE_SIZE,
                    "total": 0,
                    "total_pages": 1,
                    "date_from": date_from_str,
                    "date_to": date_to_str,
                    "client_name": client_name,
                    "client_phone_suffix": client_phone_suffix,
                    "status": status,
                    "partner_id": partner_id,
                    "partner_name": partner_name_val if 'partner_name_val' in locals() else None,
                    "to_kst_str": lambda dt: to_kst(dt).strftime("%Y-%m-%d %H:%M") if dt else "",
                    "msg": "조회 기간이 유효하지 않습니다. 다시 확인해주세요",
                },
            )

        diff_days = (d_to - d_from).days
        if diff_days > 30:  # 포함 31일 초과
            return templates.TemplateResponse(
                "partner/requests.html",
                {
                    "request": request,
                    "rows": [],
                    "page": page,
                    "page_size": PAGE_SIZE,
                    "total": 0,
                    "total_pages": 1,
                    "date_from": date_from_str,
                    "date_to": date_to_str,
                    "client_name": client_name,
                    "client_phone_suffix": client_phone_suffix,
                    "status": status,
                    "partner_id": partner_id,
                    "partner_name": partner_name_val if 'partner_name_val' in locals() else None,
                    "to_kst_str": lambda dt: to_kst(dt).strftime("%Y-%m-%d %H:%M") if dt else "",
                    "msg": "최대 조회 가능 일수는 31일입니다.",
                },
            )


    # KST → UTC 변환
    start_utc, end_utc = kst_date_range_to_utc_datetimes(d_from, d_to)

    # 기본 쿼리: 해당 파트너 매핑만
    stmt = select(PartnerClientMapping).where(
        PartnerClientMapping.partner_id == partner_id
    )

    # --- 날짜 조건: "고객신청일(client_submitted_at) OR 담당자신청일(created_at)" ---
    if start_utc or end_utc:
        def make_range(col):
            cond = None
            if start_utc:
                cond = col >= start_utc
            if end_utc:
                cond = (cond & (col < end_utc)) if cond is not None else (col < end_utc)
            return cond

        cond_client = make_range(PartnerClientMapping.client_submitted_at)
        cond_partner = make_range(PartnerClientMapping.created_at)

        combined = None
        if cond_client is not None and cond_partner is not None:
            combined = cond_client | cond_partner
        elif cond_client is not None:
            combined = cond_client
        elif cond_partner is not None:
            combined = cond_partner

        if combined is not None:
            stmt = stmt.where(combined)

    # --- 고객명: 정확 일치 ---
    if client_name:
        stmt = stmt.where(PartnerClientMapping.client_name == client_name)

    # --- 휴대폰 뒷 4자리: 정확 일치 ---
    if client_phone_suffix:
        digits = re.sub(r"[^0-9]", "", client_phone_suffix)
        if digits:
            stmt = stmt.where(func.right(PartnerClientMapping.client_phone, 4) == digits)

    # 정렬: 고객신청일(제출일) 우선, 없으면 created_at 기준
    stmt = stmt.order_by(
        PartnerClientMapping.client_submitted_at.desc(),
        PartnerClientMapping.created_at.desc(),
        PartnerClientMapping.id.desc(),
    )

    mappings = session.exec(stmt).all()

    # --- 상태 필터/리포트 발송 여부 반영하면서 rows 빌드 ---
    all_rows: list[tuple[PartnerClientMapping, Optional[Respondent], Optional[SurveyResponse], Optional[ReportFile]]] = []

    for pcm in mappings:
        resp: Optional[Respondent] = None
        sr: Optional[SurveyResponse] = None
        rf: Optional[ReportFile] = None

        # Respondent: 같은 파트너 + 이름 + 전화, 최신 1건
        try:
            if pcm.client_name and pcm.client_phone:
                resp = session.exec(
                    select(Respondent)
                    .where(
                        Respondent.partner_id == pcm.partner_id,
                        Respondent.applicant_name == pcm.client_name,
                        Respondent.client_phone == pcm.client_phone,
                    )
                    .order_by(Respondent.created_at.desc())
                ).first()
        except Exception as e:
            logging.warning(
                "[PARTNER-REQ][RESP-LOOKUP][WARN] mapping_id=%s err=%r",
                getattr(pcm, "id", None),
                e,
            )

        # ---------------------------
        # 필터: (1) 리포트 발송여부 status=unsent/sent
        # ---------------------------
        # sent: Respondent.status == report_sent
        # unsent: 그 외(응답이 없거나 report_sent가 아닌 경우 포함)
        if status == "sent":
            if not resp or resp.status != "report_sent":
                continue
        elif status == "unsent":
            if resp and resp.status == "report_sent":
                continue

        # ---------------------------
        # 필터: (2) 문진 미제출 인원만 보기 (체크박스)
        # ---------------------------
        # 문진 미제출 정의:
        # - resp.status == submitted
        # - pcm.created_at(담당자신청일)은 존재
        # - pcm.client_submitted_at(고객신청일)이 비어있음
        if only_not_submitted:
            # resp가 없어도 pcm 기준으로 문진 미제출 판단 가능해야 누락이 없음
            is_not_submitted = (pcm.created_at is not None) and (pcm.client_submitted_at is None)
            if not is_not_submitted:
                continue



        # SurveyResponse / ReportFile
        if resp:
            try:
                sr = session.exec(
                    select(SurveyResponse)
                    .where(SurveyResponse.respondent_id == resp.id)
                    .order_by(SurveyResponse.submitted_at.desc())
                ).first()
            except Exception as e:
                logging.warning(
                    "[PARTNER-REQ][SR-LOOKUP][WARN] resp_id=%s err=%r",
                    getattr(resp, "id", None),
                    e,
                )

            if sr:
                rf = session.exec(
                    select(ReportFile).where(
                        ReportFile.survey_response_id == sr.id
                    )
                ).first()

        all_rows.append((pcm, resp, sr, rf))

    # --- 페이지네이션 ---
    total = len(all_rows)
    total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE if total > 0 else 1
    if total_pages == 0:
        total_pages = 1

    if page > total_pages:
        page = total_pages

    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    rows = all_rows[start_idx:end_idx]

    # 출력용 KST 포맷
    def to_kst_str(dt: Optional[datetime]) -> str:
        return to_kst(dt).strftime("%Y-%m-%d %H:%M") if dt else ""

    # 파트너 이름(옵션)
    partner_name_val = None
    try:
        admin_user = session.get(UserAdmin, partner_id)
        if admin_user and admin_user.name:
            partner_name_val = admin_user.name
    except Exception:
        partner_name_val = None

    return templates.TemplateResponse(
        "partner/requests.html",
        {
            "request": request,
            "rows": rows,
            "page": page,
            "page_size": PAGE_SIZE,
            "total": total,
            "total_pages": total_pages,
            # 👉 템플릿에서 그대로 쓰는 날짜/필터 값들
            "date_from": date_from_str,
            "date_to": date_to_str,
            "client_name": client_name,
            "client_phone_suffix": client_phone_suffix,
            "status": status,
            "partner_id": partner_id,
            "partner_name": partner_name_val,
            "only_not_submitted": only_not_submitted,
            "to_kst_str": to_kst_str,
        },
    )

# 파트너 관리자페이지 기능
@app.get("/partner/supervisor", response_class=HTMLResponse)
def partner_supervisor(
    request: Request,
    session: Session = Depends(get_session),
    page: int = 1,
    page_size: str = "50",
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
    status: str | None = None,
    q: str | None = None,
    msg: str | None = None,
):
    # 로그인 체크
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=303)

    ua_me = session.get(UserAdmin, int(partner_id))
    if not ua_me:
        return RedirectResponse(url="/partner/login", status_code=303)

    # supervisor 권한 체크
    if not bool(getattr(ua_me, "supervisor", False)):
        return templates.TemplateResponse(
            "error.html",
            {
                "request": request,
                "message": "진입 권한이 없습니다.\n가온앤 관리자에게 문의해주세요.",
            },
            status_code=403,
        )

    # ✅ 매우 중요한 기본 조건: 같은 division(=업체명)만 노출
    my_division = (ua_me.division or "").strip()
    if not my_division:
        return templates.TemplateResponse(
            "error.html",
            {
                "request": request,
                "message": "소속(division) 정보가 없어 조회할 수 없습니다.\n가온앤 관리자에게 문의해주세요.",
            },
            status_code=400,
        )

    # 날짜 필터(기존 responses 흐름과 동일하게 KST->UTC 변환 헬퍼 사용)
    utc_from = utc_to = None
    if from_ and to:
        try:
            utc_from, utc_to = kst_date_range_to_utc_datetimes(from_, to)
        except Exception:
            utc_from = utc_to = None

    # pcm_latest: (partner_id, client_phone)별 최신 created_at (responses와 동일 패턴)
    pcm_latest = (
        select(
            PartnerClientMapping.partner_id.label("partner_id"),
            PartnerClientMapping.client_phone.label("client_phone"),
            func.max(PartnerClientMapping.created_at).label("created_at"),
        )
        .group_by(PartnerClientMapping.partner_id, PartnerClientMapping.client_phone)
        .subquery()
    )

    stmt = (
        select(
            SurveyResponse,
            Respondent,
            User,
            UserAdmin,
            pcm_latest.c.created_at.label("partner_requested_at"),
        )
        .join(Respondent, Respondent.id == SurveyResponse.respondent_id)
        .join(User, User.id == Respondent.user_id)
        .outerjoin(UserAdmin, UserAdmin.id == Respondent.partner_id)
        .outerjoin(
            pcm_latest,
            (pcm_latest.c.partner_id == Respondent.partner_id)
            & (pcm_latest.c.client_phone == Respondent.client_phone),
        )
        .where(Respondent.campaign_id == my_division)
    )

    # 문진제출일 범위
    if utc_from and utc_to:
        stmt = stmt.where(SurveyResponse.submitted_at >= utc_from, SurveyResponse.submitted_at <= utc_to)

    # 진행상태 필터 (responses와 동일 UX)
    if status:
        status = status.strip()
        if status == "analysis_requested":
            stmt = stmt.where(Respondent.status == "submitted", pcm_latest.c.created_at.is_not(None))
        elif status == "submitted":
            stmt = stmt.where(Respondent.status == "submitted", pcm_latest.c.created_at.is_(None))
        else:
            stmt = stmt.where(Respondent.status == status)

    # 텍스트 검색 (이름/생년월일/신청번호 일부)
    # --- 검색어 필터 (admin_responses와 동일 UX) ---
    stmt = _apply_supervisor_q_filter(stmt, q)

    # 정렬: 제출일 최신 우선
    stmt = stmt.order_by(SurveyResponse.submitted_at.desc())

    # 페이지 사이즈
    if page_size == "all":
        limit = None
    else:
        try:
            limit = int(page_size)
        except Exception:
            limit = 50

    # 전체 개수
    count_stmt = (
        select(func.count())
        .select_from(stmt.subquery())
    )
    total = session.exec(count_stmt).one()
    total = int(total or 0)

    # 페이징 적용
    if page < 1:
        page = 1
    if limit:
        stmt = stmt.offset((page - 1) * limit).limit(limit)

    rows = session.exec(stmt).all()

    total_pages = 1
    if limit and limit > 0:
        total_pages = (total + limit - 1) // limit

    return templates.TemplateResponse(
        "partner/supervisor.html",
        {
            "request": request,
            "rows": rows,
            "page": page,
            "total_pages": total_pages,
            "page_size": page_size,
            "from": from_ or "",
            "to": to or "",
            "status": status or "",
            "q": q or "",
            "msg": msg or "",
        },
    )


#관리자페이지 엑셀다운로드 기능
@app.get("/partner/supervisor/export.xlsx")
def partner_supervisor_export_xlsx(
    request: Request,
    session: Session = Depends(get_session),
    page: int = 1,
    page_size: str = "50",
    from_: str | None = Query(None, alias="from"),
    to: str | None = None,
    status: str | None = None,
    q: str | None = None,
):
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=303)

    ua_me = session.get(UserAdmin, int(partner_id))
    if not ua_me or not bool(getattr(ua_me, "supervisor", False)):
        return RedirectResponse(url="/partner/dashboard?msg=권한이 없습니다.", status_code=303)
    
    #로그추가 임시
    logging.info("[SVX][AUTH] ua_id=%s supervisor=%s division=%s qs=%s",
             getattr(ua_me, "id", None),
             getattr(ua_me, "supervisor", None),
             getattr(ua_me, "division", None),
             str(request.query_params))

    my_division = (ua_me.division or "").strip()
    if not my_division:
        return RedirectResponse(url="/partner/dashboard?msg=소속정보가 없습니다.", status_code=303)

    utc_from = utc_to = None
    if from_ and to:
        try:
            utc_from, utc_to = kst_date_range_to_utc_datetimes(from_, to)
        except Exception:
            utc_from = utc_to = None

    pcm_latest = (
        select(
            PartnerClientMapping.partner_id.label("partner_id"),
            PartnerClientMapping.client_phone.label("client_phone"),
            func.max(PartnerClientMapping.created_at).label("created_at"),
        )
        .group_by(PartnerClientMapping.partner_id, PartnerClientMapping.client_phone)
        .subquery()
    )

    stmt = (
        select(
            SurveyResponse,
            Respondent,
            User,
            UserAdmin,
            pcm_latest.c.created_at.label("partner_requested_at"),
        )
        .join(Respondent, Respondent.id == SurveyResponse.respondent_id)
        .join(User, User.id == Respondent.user_id)
        .outerjoin(UserAdmin, UserAdmin.id == Respondent.partner_id)
        .outerjoin(
            pcm_latest,
            (pcm_latest.c.partner_id == Respondent.partner_id)
            & (pcm_latest.c.client_phone == Respondent.client_phone),
        )
        .where(Respondent.campaign_id == my_division)
        .order_by(SurveyResponse.submitted_at.desc())
    )

    if utc_from and utc_to:
        stmt = stmt.where(SurveyResponse.submitted_at >= utc_from, SurveyResponse.submitted_at <= utc_to)

    if status:
        status = status.strip()
        if status == "analysis_requested":
            stmt = stmt.where(Respondent.status == "submitted", pcm_latest.c.created_at.is_not(None))
        elif status == "submitted":
            stmt = stmt.where(Respondent.status == "submitted", pcm_latest.c.created_at.is_(None))
        else:
            stmt = stmt.where(Respondent.status == status)

    # --- 검색어 필터 (admin_responses와 동일 UX) ---
    stmt = _apply_supervisor_q_filter(stmt, q)


    # 화면에 "보이는 테이블" 그대로 export (page/page_size 반영)
    if page_size == "all":
        limit = None
    else:
        try:
            limit = int(page_size)
        except Exception:
            limit = 50

    if page < 1:
        page = 1
    if limit:
        stmt = stmt.offset((page - 1) * limit).limit(limit)

    rows = session.exec(stmt).all()
    #로그추가 임시
    logging.info("[SVX][ROWS] division=%s rows=%s page=%s page_size=%s",
             my_division, len(rows), page, page_size)

    
    wb = Workbook()
    ws = wb.active
    ws.title = "Supervisor"

    headers = ["신청번호", "신청자", "담당자", "진행상태값", "문진제출일", "담당자신청일", "메모"]
    ws.append(headers)

    def status_label(resp_status: str, partner_requested_at) -> str:
        if resp_status == "submitted":
            return "분석신청" if partner_requested_at else "문진제출"
        if resp_status == "accepted":
            return "접수완료"
        if resp_status == "report_uploaded":
            return "리포트 업로드 완료"
        if resp_status == "report_sent":
            return "리포트 발송 완료"
        return resp_status or ""

    for sr, resp, user, ua, partner_requested_at in rows:
        applicant = (resp.applicant_name or user.name_enc or "-")
        담당자 = (ua.name if ua and ua.name else "")
        memo = (resp.sv_memo or "")

        ws.append([
            resp.serial_no or "",
            applicant,
            담당자,
            status_label(resp.status, partner_requested_at),
            to_kst(sr.submitted_at) if sr and sr.submitted_at else "",
            to_kst(partner_requested_at) if partner_requested_at else "",
            memo,
        ])

    # 열 너비(대충) + 메모는 넓게
    widths = [10, 22, 14, 14, 16, 16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    import io
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "partner_supervisor.xlsx"
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


#관리자페이지 메모 저장기능
@app.post("/partner/supervisor/memo")
def partner_supervisor_save_memo(
    request: Request,
    respondent_id: int = Form(...),
    new_memo: str = Form(""),
    next: str = Form("/partner/supervisor"),
    session: Session = Depends(get_session),
):
    partner_id = request.session.get("partner_id")
    if not partner_id:
        return RedirectResponse(url="/partner/login", status_code=303)

    ua_me = session.get(UserAdmin, int(partner_id))
    if not ua_me or not bool(getattr(ua_me, "supervisor", False)):
        return RedirectResponse(url=f"{next}?msg=권한이 없습니다.", status_code=303)

    my_division = (ua_me.division or "").strip()
    if not my_division:
        return RedirectResponse(url=f"{next}?msg=소속정보가 없습니다.", status_code=303)

    resp = session.get(Respondent, int(respondent_id))
    if not resp:
        return RedirectResponse(url=f"{next}?msg=대상이 존재하지 않습니다.", status_code=303)
    
    #로그 추가 임시
    logging.info("[SVM][TRY] ua_id=%s ua_name=%s division=%s respondent_id=%s next=%s",
             getattr(ua_me, "id", None),
             getattr(ua_me, "name", None),
             my_division,
             respondent_id,
             next)


    # ✅ 보안: 같은 campaign_id(=division)만 수정 가능
    if (resp.campaign_id or "").strip() != my_division:
        return RedirectResponse(url=f"{next}?msg=대상 접근 권한이 없습니다.", status_code=303)

    nm = (new_memo or "").strip()
    if not nm:
        return RedirectResponse(url=f"{next}?msg=신규작성 내용이 없습니다.", status_code=303)

    # 신규 입력은 60자 제한
    if len(nm) > 60:
        return RedirectResponse(url=f"{next}?msg=신규메모는 최대 60자까지 입력할 수 있습니다.", status_code=303)

    now = now_kst()
    tail = f"최종수정자: {(ua_me.name or '').strip()} / 최종수정일시: {to_kst(now)}"
    final = nm + "\n" + tail

    # DB 저장은 100자 제한 (varchar(100))
    if len(final) > 100:
        return RedirectResponse(url=f"{next}?msg=메모 저장 길이를 초과했습니다. (최대 100자)", status_code=303)

    resp.sv_memo = final
    resp.sv_memo_at = now
    resp.updated_at = now_kst()
    session.add(resp)
    session.commit()
    #저장 로그 임시
    logging.info("[SVM][SAVED] respondent_id=%s memo_len=%s memo_at=%s",
             resp.id, len(resp.sv_memo or ""), to_kst(resp.sv_memo_at) if resp.sv_memo_at else "")


    return RedirectResponse(url=f"{next}?msg=저장되었습니다.", status_code=303)



#사용자 로그인 화면 렌더
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


#로그인 코드 검증, 진행
@app.post("/login/verify")
def login_verify_phone(
    request: Request,
    phone: str = Form(...),
    session: Session = Depends(get_session),
):
    phone_digits = "".join(c for c in (phone or "") if c.isdigit())
    if len(phone_digits) < 10 or len(phone_digits) > 11:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "번호 형식이 올바르지 않습니다."},
            status_code=400,
        )

    # division까지 조회해서 campaign_id로 사용
    row = session.exec(
        sa_text("""
            SELECT id, name, phone, is_active, division
            FROM user_admin
            WHERE phone = :p AND is_active = TRUE
            LIMIT 1
        """).bindparams(p=phone_digits)
    ).first()
    if not row:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "등록되지 않은 코드입니다."},
            status_code=401,
        )

    admin_id = int(row[0])
    division = (row[4] or "").strip()
    campaign_id_value = division or DEFAULT_CAMPAIGN_ID

    # User 조회/생성
    ph = hash_phone(phone_digits)
    user = session.exec(select(User).where(User.phone_hash == ph)).first()
    if not user:
        user = User(phone_hash=ph)
        session.add(user)
        session.commit()
        session.refresh(user)

    # AUTH 쿠키 발급
    SECURE_COOKIE = bool(int(os.getenv("SECURE_COOKIE", "1")))
    resp = RedirectResponse(url="/nhis", status_code=303)
    resp.set_cookie(
        AUTH_COOKIE_NAME,
        sign_user(user.id),
        httponly=True,
        secure=SECURE_COOKIE,
        samesite="lax",
        max_age=AUTH_MAX_AGE,
    )
    resp.delete_cookie("survey_completed")

    # 세션 저장 (survey_root에서 그대로 사용)
    request.session["partner_id"] = admin_id
    request.session["admin_phone"] = phone_digits
    request.session["campaign_id"] = campaign_id_value

    # respondent를 미리 1개 생성해두고 /survey에서 재사용
    rid = session.exec(
        sa_text("""
            INSERT INTO respondent (user_id, campaign_id, status, partner_id, created_at, updated_at)
            VALUES (:uid, :cid, 'started', :pid, (now() AT TIME ZONE 'Asia/Seoul'), (now() AT TIME ZONE 'Asia/Seoul'))
            RETURNING id
        """).bindparams(uid=user.id, cid=campaign_id_value, pid=admin_id)
    ).first()[0]

    request.session["respondent_id"] = int(rid)

    # (안전) rtoken을 꼭 써야 한다면 숫자만 서명하도록
    try:
        tok = signer.sign(str(rid)).decode("utf-8")
        request.session["rtoken"] = tok
        resp.set_cookie("rtoken", tok, max_age=1800, httponly=True, samesite="Lax", secure=SECURE_COOKIE)
    except Exception as e:
        logging.debug("[LOGIN][RTOKEN][WARN] %r", e)

    logging.info("[RESP][CREATE][LOGIN] rid=%s partner_id=%s campaign_id=%s", rid, admin_id, campaign_id_value)
    return resp



@app.get("/logout")
def logout():
    resp = RedirectResponse(url="/", status_code=302)
    resp.delete_cookie(AUTH_COOKIE_NAME)
    return resp


def verify_token(token: str) -> int:
    try:
        raw = signer.unsign(token, max_age=1800*3)
        return int(raw.decode("utf-8"))
    except (BadSignature, SignatureExpired):
        return -1


@app.get("/report/ready", response_class=HTMLResponse)
def report_ready(request: Request, rtoken: str):
    return templates.TemplateResponse("report_ready.html", {"request": request, "rtoken": rtoken})

@app.get("/portal", response_class=HTMLResponse)
def portal(request: Request, auth: str | None = Cookie(default=None, alias=AUTH_COOKIE_NAME), session: Session = Depends(get_session)):
    user_id = verify_user(auth) if auth else -1
    if user_id < 0:
        return RedirectResponse(url="/login", status_code=302)

    resps = session.exec(select(Respondent).where(Respondent.user_id == user_id)).all()
    reports = []
    for r in resps:
        srs = session.exec(select(SurveyResponse).where(SurveyResponse.respondent_id == r.id)).all()
        for sr in srs:
            reports.append({
                "id": sr.id,
                "submitted_at": to_kst(sr.submitted_at).strftime("%Y-%m-%d %H:%M"),
                "score": sr.score
            })
    reports.sort(key=lambda x: x["id"], reverse=True)
    return templates.TemplateResponse("portal.html", {
        "request": request,
    })

def _normalize_date_str(s: str) -> str | None:
    """
    'YYYY-M-D' 같이 들어와도 'YYYY-MM-DD'로 0패딩해서 돌려줍니다.
    날짜 형식이 아니면 None.
    """
    s = (s or "").strip()
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if not m:
        return None
    y, mm, dd = m.groups()
    try:
        return f"{int(y):04d}-{int(mm):02d}-{int(dd):02d}"
    except ValueError:
        return None

#생년월일 날짜검색 헬퍼
def _apply_supervisor_q_filter(stmt, q: str | None):
    """
    admin_responses의 검색 UX와 동일:
    - q가 생년월일(yyyy-mm-dd / yyyymmdd 등)로 인식되면: '정확 일자 매칭(=)' + 기타 like
    - 아니면: 일반 like 검색(생년월일도 부분검색 허용)
    """
    if not q:
        return stmt

    like = f"%{q}%"
    q_birth = _normalize_date_str(q)

    if q_birth:
        # 정확 일자 매칭(=) + 기타 like
        return stmt.where(
            (func.to_char(User.birth_date, "YYYY-MM-DD") == q_birth)
            | (func.to_char(Respondent.birth_date, "YYYY-MM-DD") == q_birth)
            | (Respondent.applicant_name.ilike(like))
            | (User.name_enc.ilike(like))
            | (func.cast(Respondent.serial_no, sa.String).ilike(like))
            | (UserAdmin.name.ilike(like))
            | (Respondent.sv_memo.ilike(like))
            | (func.to_char(Respondent.created_at, "YYYY-MM-DD").ilike(like))
        )
    else:
        # 일반 텍스트 검색: 생년월일도 부분검색 허용
        return stmt.where(
            (Respondent.applicant_name.ilike(like))
            | (User.name_enc.ilike(like))
            | (func.cast(Respondent.serial_no, sa.String).ilike(like))
            | (UserAdmin.name.ilike(like))
            | (Respondent.sv_memo.ilike(like))
            | (func.to_char(User.birth_date, "YYYY-MM-DD").ilike(like))
            | (func.to_char(Respondent.birth_date, "YYYY-MM-DD").ilike(like))
            | (func.to_char(Respondent.created_at, "YYYY-MM-DD").ilike(like))
        )



#---- 관리자 로그인 ----#

def admin_required(request: Request):
    # 세션에 admin 플래그가 있으면 통과
    try:
        if request.session.get("admin"):
            return
    except Exception:
        pass
    raise HTTPException(status_code=401)

@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_form(request: Request):
    # 에러 메시지 표시용 기본값 포함
    return templates.TemplateResponse("admin/login.html", {"request": request, "error": None})


# 관리자 전용 라우터
admin_router = APIRouter(
    prefix="/admin",
    dependencies=[Depends(admin_required)]
)

@app.post("/admin/login")
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):

    # 1) ENV에서 계정 목록 읽기 (콤마 지원)
    #    예: ADMIN_USER="admin1,admin2"  ADMIN_PASS="pass1,pass2"
    env_users = [u.strip() for u in (os.getenv("ADMIN_USER") or "").replace("\n", "").split(",") if u.strip()]
    env_pwds  = [p.strip() for p in (os.getenv("ADMIN_PASS") or "").replace("\n", "").split(",") if p.strip()]

    # 2) 방어: 개수 불일치 시 뒤쪽 잘라내기
    n = min(len(env_users), len(env_pwds))
    env_users = env_users[:n]
    env_pwds  = env_pwds[:n]

    # 3) 1:1 인덱스 매칭으로 검증
    valid = any((username == env_users[i] and password == env_pwds[i]) for i in range(n))

    if not valid:
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "message": "인증 실패"},
            status_code=401,
        )

    # 4) 세션 발급 (기존 키 그대로)
    # admin 세션만 설정 (partner 세션과 공존)
    request.session["admin"] = True
    request.session["_iat"] = int(datetime.now(timezone.utc).timestamp())

    return RedirectResponse(url="/admin/responses", status_code=303)


@app.get("/admin/logout")
def admin_logout(request: Request):
    # admin 세션만 제거 (partner 세션 유지)
    request.session.pop("admin", None)
    request.session.pop("_iat", None)

    return RedirectResponse(url="/admin/login", status_code=303)


# 목록
@admin_router.get("/responses", response_class=HTMLResponse)
def admin_responses(
    request: Request,
    response: Response,
    page: int = 1,
    page_size: str = "50",  # "50"(기본) / "100" / "all"
    q: Optional[str] = None,
    status: Optional[str] = None,  # submitted/accepted/report_uploaded or ""
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = None,
    msg: Optional[str] = None,
    session: Session = Depends(get_session),
):
    
    # --- 첫 진입 기본값: from/to가 없으면 오늘(KST)로 세팅 ---
    is_first_visit = (request.url.query == "")
    today_str = now_kst().date().isoformat()
    
    # 안전 파싱
    try:
        page = max(1, int(page))
    except Exception:
        page = 1

    # --- page_size 해석 ---
    page_size_norm = (page_size or "50").lower()
    if page_size_norm == "100":
        PAGE_SIZE = 100
    elif page_size_norm == "all":
        PAGE_SIZE = None  # 전체 보기
    else:
        PAGE_SIZE = 50

    # --- 기본 쿼리 ---
    # 담당자(UserAdmin) + 담당자신청일(PartnerClientMapping.created_at)을 함께 조회
    # PartnerClientMapping은 (partner_id, client_phone)별 최신 1건만 붙여서 row 중복(뻥튀기)을 방지
    pcm_latest = (
        select(
            PartnerClientMapping.partner_id.label("partner_id"),
            PartnerClientMapping.client_phone.label("client_phone"),
            func.max(PartnerClientMapping.created_at).label("created_at"),
        )
        .group_by(PartnerClientMapping.partner_id, PartnerClientMapping.client_phone)
        .subquery()
    )

    stmt = (
        select(
            SurveyResponse,
            Respondent,
            User,
            ReportFile,
            UserAdmin,
            pcm_latest.c.created_at.label("partner_requested_at"),
        )
        .join(Respondent, Respondent.id == SurveyResponse.respondent_id)
        .join(User, User.id == Respondent.user_id)
        .join(ReportFile, ReportFile.survey_response_id == SurveyResponse.id, isouter=True)
        .join(UserAdmin, UserAdmin.id == Respondent.partner_id, isouter=True)
        .join(
            pcm_latest,
            and_(
                pcm_latest.c.partner_id == Respondent.partner_id,
                pcm_latest.c.client_phone == Respondent.client_phone,
            ),
            isouter=True,
        )
    )



    # --- 검색어 필터 (생년월일 yyyy-mm-dd 지원) ---
    if q:
        like = f"%{q}%"
        q_birth = _normalize_date_str(q)

        if q_birth:
            # 정확한 일자 매칭(=) + 기존 like들
            stmt = stmt.where(
                (func.to_char(User.birth_date, "YYYY-MM-DD") == q_birth)
                | (func.to_char(Respondent.birth_date, "YYYY-MM-DD") == q_birth)
                | (User.name_enc.ilike(like))
                | (User.phone_hash.ilike(like))
                | (SurveyResponse.answers_json.ilike(like))
                | (ReportFile.filename.ilike(like))
                | (func.to_char(Respondent.created_at, "YYYY-MM-DD").ilike(like))
            )
        else:
            # 일반 텍스트 검색: 생년월일도 부분검색 허용
            stmt = stmt.where(
                (User.name_enc.ilike(like))
                | (User.phone_hash.ilike(like))
                | (SurveyResponse.answers_json.ilike(like))
                | (ReportFile.filename.ilike(like))
                | (func.to_char(Respondent.created_at, "YYYY-MM-DD").ilike(like))
                | (func.to_char(User.birth_date, "YYYY-MM-DD").ilike(like))
                | (func.to_char(Respondent.birth_date, "YYYY-MM-DD").ilike(like))
            )


    # --- 날짜 필터 ---
    def parse_date(s: Optional[str]):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    d_from = parse_date(from_)
    d_to = parse_date(to)
    # 기간 검사 (둘 다 있는 경우만)
    if d_from and d_to:
        if d_from > d_to:
            return _redirect_with_msg(request.url.path + (("?" + request.url.query) if request.url.query else ""), "조회 기간이 유효하지 않습니다. 다시 확인해주세요")
        diff_days = (d_to - d_from).days
        if diff_days > 30:  # 포함 31일 초과
            return _redirect_with_msg(request.url.path + (("?" + request.url.query) if request.url.query else ""), "최대 조회 가능 일수는 31일입니다.")

    start_utc, end_utc = kst_date_range_to_utc_datetimes(d_from, d_to)
    # 문진제출일(= SurveyResponse.submitted_at) 기준으로 필터
    if start_utc:
        stmt = stmt.where(SurveyResponse.submitted_at >= start_utc)
    if end_utc:
        stmt = stmt.where(SurveyResponse.submitted_at < end_utc)

    # --- 상태 필터 ---
    # - 문진제출: Respondent.status == submitted AND 담당자신청일(created_at)이 없음
    # - 분석신청: Respondent.status == submitted AND 담당자신청일(created_at)이 있음
    # - 접수완료/리포트업로드 완료: 기존 status 그대로
    if status == "analysis_requested":
        stmt = stmt.where(
            (Respondent.status == "submitted") &
            (pcm_latest.c.created_at.isnot(None))
        )
    elif status == "submitted":
        stmt = stmt.where(
            (Respondent.status == "submitted") &
            (pcm_latest.c.created_at.is_(None))
        )
    elif status in ("accepted", "report_uploaded", "report_sent"):
        stmt = stmt.where(Respondent.status == status)


    # --- 정렬: 최신 제출일 → 응답 ID 내림차순 ---
    stmt = stmt.order_by(
        SurveyResponse.submitted_at.desc(),
        SurveyResponse.id.desc()
    )

    # --- 전체 개수 ---
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = session.exec(count_stmt).one()

    # --- 페이징 ---
    if PAGE_SIZE is None:
        # 전체보기: offset/limit 제거
        rows = session.exec(stmt).all()
        total_pages = 1
        page = 1
    else:
        rows = session.exec(
            stmt.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE)
        ).all()
        total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE

    # --- 출력 변환 ---
    to_kst_str = lambda dt: to_kst(dt).strftime("%Y-%m-%d %H:%M")


    # --- 렌더 ---
    return templates.TemplateResponse(
        "admin/responses.html",
        {
            "request": request,
            "rows": rows,
            "page": page,
            "total": total,
            "total_pages": total_pages,
            "from": from_,
            "to": to,
            "status": status or "",
            "q": q,
            "msg": msg or "",
            "page_size": page_size_norm,
            "to_kst_str": to_kst_str,
        },
    )

#responses 페이지 안전 리다이렉트 헬퍼
def _safe_next_url(next_url: str | None) -> str:
    if not next_url:
        return "/admin/responses"
    # 외부 URL 오픈리다이렉트 방지: 내부 경로만 허용
    if next_url.startswith("/admin/responses"):
        return next_url
    if next_url.startswith("/admin/response"):
        # 혹시 개별 액션 후 responses로 돌리고 싶다면 제한적으로 허용 가능
        return "/admin/responses"
    return "/admin/responses"

#리포트 발송 안내 문자열에서 "기호 제거
def _redirect_with_msg(next_url: str | None, msg: str) -> RedirectResponse:
    url = _safe_next_url(next_url)
    if msg:
        url = url + ("&" if "?" in url else "?") + "msg=" + urllib.parse.quote(msg)
    return RedirectResponse(url=url, status_code=303)


# 접수완료 처리 (POST + Form)
@admin_router.post("/responses/accept")
async def admin_bulk_accept(
    request: Request,
    response: Response,
    next: str | None = Form(None),
    ids: str = Form(...),  # "1,2,3"
    session: Session = Depends(get_session),
):

    # ... 나머지 처리 및 RedirectResponse 반환
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return RedirectResponse(url=_safe_next_url(next), status_code=303)
    srs = session.exec(select(SurveyResponse).where(SurveyResponse.id.in_(id_list))).all()
    sent_cnt = 0
    accepted_cnt = 0

    for sr in srs:
        resp = session.get(Respondent, sr.respondent_id)
        if not resp:
            continue
        if resp.status == "report_sent":
            sent_cnt += 1
            continue
        resp.status = "accepted"
        session.add(resp)
        accepted_cnt += 1

    session.commit()

    if sent_cnt > 0:
        msg = (
            f"이미 리포트를 발송한 건 {sent_cnt}건이 포함되어있습니다.\n"
            f"해당 건을 제외하고 접수완료 처리를 진행하였습니다.\n"
            f"(접수완료 처리 {accepted_cnt}건)"
        )
        return _redirect_with_msg(next, msg)


@admin_router.post("/responses/report/send")
async def admin_send_reports(
    ids: str = Form(...),
    next: str | None = Form(None),
    session: Session = Depends(get_session),
):
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return RedirectResponse(url=_safe_next_url(next), status_code=303)

    # 담당자신청일(PartnerClientMapping.created_at) 최신 1건만 (row 중복 방지)
    pcm_latest = (
        select(
            PartnerClientMapping.partner_id.label("partner_id"),
            PartnerClientMapping.client_phone.label("client_phone"),
            func.max(PartnerClientMapping.created_at).label("created_at"),
        )
        .group_by(PartnerClientMapping.partner_id, PartnerClientMapping.client_phone)
        .subquery()
    )

    stmt = (
        select(
            SurveyResponse,
            Respondent,
            User,
            ReportFile,
            UserAdmin,
            pcm_latest.c.created_at.label("partner_requested_at"),
        )
        .join(Respondent, Respondent.id == SurveyResponse.respondent_id)
        .join(User, User.id == Respondent.user_id)
        .join(ReportFile, ReportFile.survey_response_id == SurveyResponse.id, isouter=True)
        .join(UserAdmin, UserAdmin.id == Respondent.partner_id, isouter=True)
        .join(
            pcm_latest,
            and_(
                pcm_latest.c.partner_id == Respondent.partner_id,
                pcm_latest.c.client_phone == Respondent.client_phone,
            ),
            isouter=True,
        )
        .where(SurveyResponse.id.in_(id_list))
    )

    rows = session.exec(stmt).all()
    
    # DEBUG: 어떤 데이터가 들어왔는지 즉시 확인
    print("[REPORT-SEND] rows=", len(rows))
    for sr, resp, user, rf, ua, partner_requested_at in rows:
        print(
            "[REPORT-SEND] sr.id=", sr.id,
            " resp.id=", resp.id if resp else None,
            " status=", resp.status if resp else None,
            " to_email=", (ua.mail if ua else None),
            " rf=", bool(rf),
            " rf.content.len=", (len(rf.content) if (rf and getattr(rf, "content", None)) else None),
    )

    if not rows:
        return RedirectResponse(url=_safe_next_url(next), status_code=303)

    already_sent = 0
    not_uploaded = 0
    missing_email = 0
    missing_report = 0

    # 1) 선검증: 하나라도 조건 안 맞으면 "상태 변경/메일발송" 절대 하지 않고 즉시 중단
    for sr, resp, user, rf, ua, partner_requested_at in rows:
        if resp.status == "report_sent":
            already_sent += 1
            continue
        if resp.status != "report_uploaded":
            not_uploaded += 1
            continue
        if not rf or not getattr(rf, "content", None):
            missing_report += 1
            continue
        if not ua or not (ua.mail or "").strip():
            missing_email += 1
            continue

    if already_sent > 0:
        msg = (
            f"이미 리포트 발송을 완료한 {already_sent}건이 포함되어있습니다.\n"
            f"확인 후 다시 시도해주세요."
        )
        return _redirect_with_msg(next, msg)

    if not_uploaded > 0:
        msg = (
            f"선택한 목록 중 아직 리포트 업로드가 완료되지 않은 건이 {not_uploaded}건 있습니다.\n"
            f"확인 후 다시 시도해주세요."
        )
        return _redirect_with_msg(next, msg)

    if missing_report > 0:
        msg = (
            f"선택한 목록 중 리포트 파일이 없는 건이 {missing_report}건 있습니다.\n"
            f"확인 후 다시 시도해주세요."
        )
        return _redirect_with_msg(next, msg)

    if missing_email > 0:
        msg = (
            f"선택한 목록 중 담당자 메일 정보가 없는 건이 {missing_email}건 있습니다.\n"
            f"확인 후 다시 시도해주세요."
        )
        return _redirect_with_msg(next, msg)

    # 2) 실제 발송: 여기까지 왔으면 모두 report_uploaded + 메일/파일 보장
    ok_cnt = 0
    fail_cnt = 0

    for sr, resp, user, rf, ua, partner_requested_at in rows:
        # 제목: 신청자명 2번째 글자 마스킹
        applicant_name_raw = (resp.applicant_name or user.name_enc or "").strip()
        masked_applicant = mask_second_char(applicant_name_raw) if applicant_name_raw else ""

        partner_name = (ua.name or "").strip()
        to_email = (ua.mail or "").strip()

        # 담당자신청일(KST 문자열)
        partner_requested_at_kst_str = ""
        if partner_requested_at:
            try:
                partner_requested_at_kst_str = to_kst(partner_requested_at).strftime("%Y-%m-%d")
            except Exception:
                partner_requested_at_kst_str = ""

        ok = send_report_email(
            to_email=to_email,
            partner_name=partner_name,
            applicant_name=applicant_name_raw,  # 본문은 마스킹 불필요(요청사항)
            partner_requested_at_kst_str=partner_requested_at_kst_str,
            pdf_filename=(rf.filename or "report.pdf"),
            pdf_bytes=rf.content,
        )

        if ok:
            resp.status = "report_sent"
            resp.report_sent_at = now_kst()
            session.add(resp)
            ok_cnt += 1
        else:
            fail_cnt += 1

    session.commit()

    if fail_cnt > 0:
        msg = f"리포트 발송 완료 {ok_cnt}건, 실패 {fail_cnt}건이 있습니다.\n로그를 확인해주세요."
        return _redirect_with_msg(next, msg)

    msg = f"리포트 발송을 완료했습니다. ({ok_cnt}건)"
    return _redirect_with_msg(next, msg)



# 리포트 업로드 (POST + multipart/form-data)
@admin_router.post("/response/{rid}/report")
async def admin_upload_report(
    rid: int,
    file: UploadFile = File(...),
    next: str | None = Form(None),
    session: Session = Depends(get_session),
):
    sr = session.get(SurveyResponse, rid)
    if not sr:
        raise HTTPException(status_code=404, detail="not found")
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF만 업로드 가능합니다.")
    content = await file.read()

    # 기존 파일 있으면 교체
    old = session.exec(select(ReportFile).where(ReportFile.survey_response_id == rid)).first()
    if old:
        session.delete(old); session.commit()

    rf = ReportFile(survey_response_id=rid, filename=file.filename, content=content)
    session.add(rf)

    # 상태 업데이트
    with session.no_autoflush:
        resp = session.get(Respondent, sr.respondent_id)
    if resp:
        resp.status = "report_uploaded"
        resp.report_sent_at = None
        session.add(resp)

    session.commit()
    return RedirectResponse(url=_safe_next_url(next), status_code=303)


# 리포트 삭제 (POST)
@admin_router.post("/response/{rid}/report/delete")
def admin_delete_report(
    rid: int,
    next: str | None = Form(None),
    session: Session = Depends(get_session),
):
    sr = session.get(SurveyResponse, rid)
    if not sr:
        raise HTTPException(status_code=404, detail="not found")
    rf = session.exec(select(ReportFile).where(ReportFile.survey_response_id == rid)).first()
    if rf:
        session.delete(rf)
    # 상태는 업로드 이전 단계로(접수완료 유지)
    resp = session.get(Respondent, sr.respondent_id)
    if resp and resp.status == "report_uploaded":
        resp.status = "accepted"
        session.add(resp)
    session.commit()
    return RedirectResponse(url=_safe_next_url(next), status_code=303)

#리포트 다운로드
@admin_router.get("/response/{rid}/report/download")
def admin_download_report(rid: int, session: Session = Depends(get_session)):
    rf = session.exec(
        select(ReportFile).where(ReportFile.survey_response_id == rid)
    ).first()

    if not rf or not getattr(rf, "content", None):
        # JSON 404 대신, 관리자 화면으로 msg 리다이렉트
        return _redirect_with_msg("/admin/responses", "리포트 파일이 존재하지 않습니다.")

    filename = (rf.filename or f"report_{rid}.pdf").strip()
    if not filename.lower().endswith(".pdf"):
        filename += ".pdf"

    # UTF-8 파일명 대응 (공백/괄호 등)
    quoted = urllib.parse.quote(filename)

    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quoted}"
    }

    return StreamingResponse(
        BytesIO(rf.content),
        media_type="application/pdf",
        headers=headers
    )

# CSV (GET)
@admin_router.get("/responses.csv")
def admin_responses_csv(
    q: Optional[str] = None,
    min_score: Optional[str] = None,
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None, alias="to"),
    status: Optional[str] = None,
    session: Session = Depends(get_session),
):
    def parse_date(s: Optional[str]):
        try: return datetime.strptime(s, "%Y-%m-%d").date()
        except: return None
    d_from, d_to = parse_date(from_), parse_date(to)

    def to_kst_str(dt):
        return to_kst(dt).strftime("%Y-%m-%d %H:%M") if dt else ""

    # 담당자신청일(PartnerClientMapping.created_at) 최신 1건만 (row 중복 방지)
    pcm_latest = (
        select(
            PartnerClientMapping.partner_id.label("partner_id"),
            PartnerClientMapping.client_phone.label("client_phone"),
            func.max(PartnerClientMapping.created_at).label("created_at"),
        )
        .group_by(PartnerClientMapping.partner_id, PartnerClientMapping.client_phone)
        .subquery()
    )

    stmt = (
        select(
            SurveyResponse,
            Respondent,
            User,
            ReportFile,
            UserAdmin,
            pcm_latest.c.created_at.label("partner_requested_at"),
        )
        .join(Respondent, Respondent.id == SurveyResponse.respondent_id)
        .join(User, User.id == Respondent.user_id)
        .join(ReportFile, ReportFile.survey_response_id == SurveyResponse.id, isouter=True)
        .join(UserAdmin, UserAdmin.id == Respondent.partner_id, isouter=True)
        .join(
            pcm_latest,
            and_(
                pcm_latest.c.partner_id == Respondent.partner_id,
                pcm_latest.c.client_phone == Respondent.client_phone,
            ),
            isouter=True,
        )
    )


    if q:
        like = f"%{q}%"
        q_birth = _normalize_date_str(q)

        if q_birth:
            stmt = stmt.where(
                (func.to_char(User.birth_date, "YYYY-MM-DD") == q_birth)
                | (func.to_char(Respondent.birth_date, "YYYY-MM-DD") == q_birth)
                | (User.name_enc.ilike(like))
                | (User.phone_hash.ilike(like))
                | (SurveyResponse.answers_json.ilike(like))
                | (ReportFile.filename.ilike(like))
                | (func.to_char(Respondent.created_at, "YYYY-MM-DD").ilike(like))
            )
        else:
            stmt = stmt.where(
                (User.name_enc.ilike(like))
                | (User.phone_hash.ilike(like))
                | (SurveyResponse.answers_json.ilike(like))
                | (ReportFile.filename.ilike(like))
                | (func.to_char(Respondent.created_at, "YYYY-MM-DD").ilike(like))
                | (func.to_char(User.birth_date, "YYYY-MM-DD").ilike(like))
                | (func.to_char(Respondent.birth_date, "YYYY-MM-DD").ilike(like))
            )


    try:
        min_score_i = int(min_score) if (min_score not in (None, "")) else None
    except ValueError:
        min_score_i = None
    if min_score_i is not None:
        stmt = stmt.where(SurveyResponse.score >= min_score_i)

    # 날짜 필터(KST→UTC)
    start_utc, end_utc = kst_date_range_to_utc_datetimes(d_from, d_to)
    if start_utc:
        stmt = stmt.where(Respondent.created_at >= start_utc)
    if end_utc:
        stmt = stmt.where(Respondent.created_at < end_utc)

    if status == "analysis_requested":
        stmt = stmt.where(
            (Respondent.status == "submitted") &
            (pcm_latest.c.created_at.isnot(None))
        )
    elif status == "submitted":
        stmt = stmt.where(
            (Respondent.status == "submitted") &
            (pcm_latest.c.created_at.is_(None))
        )
    elif status in ("accepted", "report_uploaded"):
        stmt = stmt.where(Respondent.status == status)


    def generate():
        yield "\ufeff"
        s = StringIO(); w = csv.writer(s)
        w.writerow(["신청번호","신청자","담당자","담당자 소속","PDF 파일명","업로드 날짜","진행 상태값","문진제출일","담당자신청일"])
        yield s.getvalue(); s.seek(0); s.truncate(0)

        result = session.exec(stmt.order_by(SurveyResponse.submitted_at.desc())).all()
        for idx, (sr, resp, user, rf, ua, partner_requested_at) in enumerate(result, start=1):
            applicant = f"{resp.applicant_name or (user.name_enc or '')} ({(resp.birth_date or '')}, {resp.gender or (user.gender or '')})"

            if resp.status == "submitted":
                status_h = "분석신청" if partner_requested_at else "문진제출"
            elif resp.status == "accepted":
                status_h = "접수완료"
            elif resp.status == "report_uploaded":
                status_h = "리포트 업로드 완료"
            else:
                status_h = (resp.status or "")

            row = [
                resp.serial_no or "",
                applicant,
                (ua.name if ua else ""),
                (ua.division if ua else ""),
                (rf.filename if rf else ""),
                (to_kst_str(rf.uploaded_at) if rf else ""),
                status_h,
                to_kst_str(sr.submitted_at),
                to_kst_str(partner_requested_at),
            ]
            w.writerow(row)
            yield s.getvalue(); s.seek(0); s.truncate(0)


    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="responses.csv"'}
    )


# 개별 응답 2행 CSV (GET)
@admin_router.get("/response/{rid}.csv")
def admin_response_two_rows_csv(
    rid: int,
    session: Session = Depends(get_session),
):
    sr = session.get(SurveyResponse, rid)
    if not sr:
        raise HTTPException(status_code=404, detail="not found")

    try:
        payload = json.loads(sr.answers_json) if sr.answers_json else {}
    except Exception:
        payload = {}

    answers = payload.get("answers_indices") or []

    # 질문 타이틀 추출(질문 id 오름차순, 키 다양성 대응)
    def q_title(q: dict):
        return q.get("title") or q.get("text") or q.get("label") or q.get("question") or q.get("prompt") or q.get("name") or f"Q{q.get('id','')}"
    questions = [q_title(q) for q in sorted(ALL_QUESTIONS, key=lambda x: x["id"])]

    def fmt(val):
        if isinstance(val, list):
            return ",".join(str(v) for v in val)
        return "" if val is None else str(val)

    answer_numbers = [fmt(v) for v in answers]

    sio = StringIO()
    w = csv.writer(sio)
    w.writerow(questions)
    w.writerow(answer_numbers)
    csv_bytes = sio.getvalue().encode("utf-8-sig")

    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename=\"response_{rid}.csv\"'},
    )

app.include_router(admin_router)

#---- 문진 가져오기 ----#
@app.get("/survey")
def survey_root(
    request: Request,
    auth: str | None = Cookie(default=None, alias=AUTH_COOKIE_NAME),
    session: Session = Depends(get_session),
):
    user_id = verify_user(auth) if auth else -1
    if user_id < 0:
        return RedirectResponse(url="/login", status_code=302)

    user = session.get(User, user_id)

    has_birth = bool(getattr(user, "birth_date", None) or getattr(user, "birth_year", None))
    if not user or not user.name_enc or not user.gender or not has_birth:
        return RedirectResponse(url="/info", status_code=303)

    try:
        sess = request.session or {}
    except Exception:
        sess = {}

    campaign_id_value = (str(sess.get("campaign_id") or "").strip() or DEFAULT_CAMPAIGN_ID)

    # partner_id 결정
    partner_id_value: int | None = None
    raw_pid = sess.get("partner_id")
    if raw_pid is not None:
        try:
            partner_id_value = int(raw_pid)
        except (TypeError, ValueError):
            partner_id_value = None

    if partner_id_value is None:
        admin_phone = sess.get("admin_phone")
        if admin_phone:
            row = session.exec(
                sa_text("""
                    SELECT id
                      FROM user_admin
                     WHERE phone = :p
                       AND is_active = TRUE
                     LIMIT 1
                """).bindparams(p=admin_phone)
            ).first()
            if row:
                partner_id_value = int(row[0])

    # ✅ login/캠페인 start에서 미리 만든 respondent 재사용
    resp: Respondent | None = None
    existing_rid = sess.get("respondent_id")
    if existing_rid is not None:
        try:
            rid_int = int(existing_rid)
            cand = session.get(Respondent, rid_int)
            if cand and cand.user_id == user.id and cand.status != "submitted":
                resp = cand
        except Exception:
            resp = None

    if resp:
        resp.campaign_id = campaign_id_value
        resp.partner_id = partner_id_value
        if resp.status == "started":
            resp.status = "draft"
        resp.updated_at = now_kst()
    else:
        resp = Respondent(
            user_id=user.id,
            campaign_id=campaign_id_value,
            status="draft",
            partner_id=partner_id_value,
        )
        session.add(resp)
        session.commit()
        session.refresh(resp)
        request.session["respondent_id"] = resp.id

    # User 스냅샷
    bd = getattr(user, "birth_date", None)
    if not bd and getattr(user, "birth_year", None):
        bd = date(user.birth_year, 1, 1)

    resp.applicant_name = user.name_enc
    resp.birth_date = bd
    resp.gender = user.gender

    session.add(resp)
    session.commit()

    rtoken = signer.sign(str(resp.id)).decode("utf-8")
    redirect = RedirectResponse(url=f"/survey/step/1?rtoken={rtoken}", status_code=303)
    redirect.delete_cookie("survey_completed")
    return redirect



@app.get("/survey/step/{step}", response_class=HTMLResponse)
def survey_step_get(
    request: Request,
    step: int,
    rtoken: str,
    acc: str | None = None,
    _guard: None = Depends(ensure_not_completed),
    session: Session = Depends(get_session),
):
    if step < 1 or step > 3:
        return RedirectResponse(url="/survey/step/1", status_code=303)

    respondent_id = verify_token(rtoken)
    if respondent_id < 0:
        return RedirectResponse(url="/login", status_code=302)
    
    # ★ 이미 제출된 respondent면 접근 막기 (다른 브라우저/기기에서도)
    resp = session.get(Respondent, respondent_id)
    if not resp or resp.status == "submitted":
        return RedirectResponse(url="/", status_code=302)

    # 여기서 step별 문항은 헬퍼 함수에서 처리
    questions = get_questions_for_step(step)

    return templates.TemplateResponse(
        "survey_page.html",
        {
            "request": request,
            "step": step,
            "questions": questions,
            "acc": acc or "{}",
            "rtoken": rtoken,
            "is_last": step == 3,
            "is_first": step == 1,
        },
    )


        
@app.post("/survey/step/{step}")
async def survey_step_post(request: Request, step: int,
                           acc: str = Form("{}"),
                           rtoken: str = Form(...)):
    respondent_id = verify_token(rtoken)
    if respondent_id < 0:
        return RedirectResponse(url="/login", status_code=302)

    form = await request.form()

    try:
        acc_data = json.loads(acc) if acc else {}
    except Exception:
        acc_data = {}

    for q in get_questions_for_step(step):
        key = f"q{q['id']}"
        if q["type"] == "checkbox":
            vals = form.getlist(key)
            if vals:
                acc_data[key] = vals
        else:
            val = form.get(key)
            if val:
                acc_data[key] = val

    acc_q = json.dumps(acc_data, ensure_ascii=False)
    if step < 3:
        return RedirectResponse(
            url=f"/survey/step/{step+1}?acc={acc_q}&rtoken={rtoken}",
            status_code=303
        )
    else:
        return RedirectResponse(
            url=f"/survey/finish?acc={acc_q}&rtoken={rtoken}",   # ← 제출 엔드포인트 변경
            status_code=303
        )


@app.get("/survey/finish")
def survey_finish(
    request: Request,
    acc: str,
    rtoken: str,
    background_tasks: BackgroundTasks,
    session: Session = Depends(get_session),
):
    """문진 제출 처리 및 NHIS 검진 데이터 저장"""
    respondent_id = verify_token(rtoken)
    if respondent_id < 0:
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "message": "세션이 만료되었습니다. 다시 시작해주세요."},
            status_code=401,
        )

    # === NHIS 최종 수집: 세션의 '작은 값'(picked_tmp) + nhis_audit 원문(raw_from_audit) ===
    try:
        picked_tmp = (request.session or {}).get("nhis_latest") or {}

        # 1순위: 이번 세션의 callbackId로 감사로그에서 가장 최근 원문
        cbid = (request.session or {}).get("nhis_callback_id")
        raw_from_audit = None
        if cbid:
            row = session.exec(sa_text("""
                SELECT response_json
                  FROM nhis_audit
                 WHERE callback_id = :cbid
                 ORDER BY id DESC
                 LIMIT 1
            """).bindparams(cbid=cbid)).first()
            if row:
                raw_from_audit = row[0]

        # 2순위: respondent_id로 최근 감사로그
        if raw_from_audit is None:
            row2 = session.exec(sa_text("""
                SELECT response_json
                  FROM nhis_audit
                 WHERE respondent_id = :rid
                 ORDER BY id DESC
                 LIMIT 1
            """).bindparams(rid=respondent_id)).first()
            if row2:
                raw_from_audit = row2[0]

        # 표준값이 비어있고 원문이 있으면, 원문으로부터 '최근 1건' 표준화 생성
        if (not picked_tmp) and raw_from_audit:
            try:
                picked_tmp = pick_latest_general(raw_from_audit, mode="latest")
            except Exception:
                picked_tmp = {}

        # 이후 로직에서 사용할 이름으로 통일
        nhis_latest = picked_tmp or {}
        nhis_raw    = raw_from_audit or {}

    except Exception as e:
        logging.error("[NHIS][FINISH][ERR-prep] %r", e)
        nhis_latest, nhis_raw = {}, {}

    # ────────────────────────────────────────────────────────────────
    # 응답 폼(acc) 파싱
    try:
        acc_obj = json.loads(acc) if acc else {}
    except Exception:
        acc_obj = {}

    # 정규화된 답안 인덱스 구성
    answers_indices = []
    for q in sorted(ALL_QUESTIONS, key=lambda x: x["id"]):
        key = f"q{q['id']}"
        if q.get("type") == "checkbox":
            raw_list = acc_obj.get(key, [])
            if isinstance(raw_list, str):
                raw_list = [s.strip() for s in raw_list.split(",") if s.strip()]
            idx_list = []
            for v in raw_list:
                try:
                    i = int(v)
                    if i >= 1:
                        idx_list.append(i)
                except:
                    pass
            answers_indices.append(idx_list)
        else:
            v = acc_obj.get(key)
            try:
                i = int(v) if v not in (None, "") else None
                answers_indices.append(i if (i is None or i >= 1) else None)
            except:
                answers_indices.append(None)

    normalized_payload = {"answers_indices": answers_indices}
    if nhis_latest:
        normalized_payload["nhis"] = nhis_latest

    # 응답자/사용자 조회 (인적사항)
    resp = session.get(Respondent, respondent_id)
    user = session.get(User, resp.user_id) if resp and resp.user_id else None
    
    
    # ── 간편인증 시점의 동의 정보(세션) → 이번 설문 respondent에 최종 반영 ──
    try:
        sess = request.session or {}
        consent_all = bool(sess.get("agreement_all"))
        consent_at_str = sess.get("agreement_at")  # now_kst().isoformat() 형태 문자열

        if resp and consent_all:
            # 이미 true면 다시 false로 바꾸지 않음
            if not getattr(resp, "agreement_all", False):
                resp.agreement_all = True

            # 동의 시각이 비어있으면 세션에 남겨둔 시각을 쓰고, 없으면 지금 시각
            if getattr(resp, "agreement_at", None) is None:
                if consent_at_str:
                    try:
                        at_val = datetime.fromisoformat(consent_at_str)
                    except Exception:
                        at_val = now_kst()
                else:
                    at_val = now_kst()
                resp.agreement_at = at_val

            resp.updated_at = now_kst()
            session.add(resp)
            session.commit()
    except Exception as e:
        logging.warning("[CONSENT][WARN][finish] apply failed: %r", e)

    
    # NHIS 세션 정보 → Respondent에 반영 (고객 이름/휴대폰)
    if resp:
        sync_respondent_contact_from_nhis(request, session, resp)


        # 동의 여부 세션 → Respondent 컬럼으로 반영
    if resp:
        try:
            sess = request.session or {}
            agr_all = bool(sess.get("agreement_all"))
            agr_at_str = sess.get("agreement_at")

            if agr_all:
                resp.agreement_all = True
                # 이미 값이 있으면 덮어쓰지 않고, 없을 때만 기록
                if not resp.agreement_at:
                    try:
                        # 세션에 isoformat()으로 넣어둔 값을 되살림
                        resp.agreement_at = datetime.fromisoformat(agr_at_str) if agr_at_str else now_kst()
                    except Exception:
                        resp.agreement_at = now_kst()
        except Exception as e:
            logging.warning("[CONSENT][WARN] agreement sync failed: %r", e)


    # 로그인 시점 정보로 partner_id 보정
    if resp and not resp.partner_id:
        try:
            sess = request.session or {}
        except Exception:
            sess = {}

        # 1) 세션의 partner_id 먼저 시도
        pid_from_session = sess.get("partner_id")
        partner_id_value: int | None = None
        if pid_from_session is not None:
            try:
                partner_id_value = int(pid_from_session)
            except (TypeError, ValueError):
                partner_id_value = None

        # 2) 없으면 admin_phone → user_admin.id 조회
        if partner_id_value is None:
            admin_phone = sess.get("admin_phone")
            if admin_phone:
                try:
                    row = session.exec(
                        sa_text("""
                            SELECT id
                              FROM user_admin
                             WHERE phone = :p
                               AND is_active = TRUE
                             LIMIT 1
                        """).bindparams(p=admin_phone)
                    ).first()
                    if row:
                        partner_id_value = int(row[0])
                        logging.info(
                            "[RESP][FIX-PID][ADMIN-PHONE] resp_id=%s admin_phone=%s partner_id=%s",
                            resp.id,
                            admin_phone,
                            partner_id_value,
                        )
                except Exception as e:
                    logging.warning(
                        "[RESP][FIX-PID][LOOKUP-FAIL] resp_id=%s phone=%s err=%r",
                        resp.id,
                        admin_phone,
                        e,
                    )

        if partner_id_value is not None:
            resp.partner_id = partner_id_value
            session.add(resp)
            session.commit()
            logging.info(
                "[RESP][FIX-PID] resp_id=%s partner_id=%s",
                resp.id,
                resp.partner_id,
            )


    def calc_age(bd, ref_date):
        if not bd:
            return ""
        return ref_date.year - bd.year - ((ref_date.month, ref_date.day) < (bd.month, bd.day))

    today = now_kst().date()
    name = (resp.applicant_name if resp and resp.applicant_name else (user.name_enc if user and user.name_enc else "")) or ""
    bd = resp.birth_date if (resp and resp.birth_date) else (getattr(user, "birth_date", None) if user else None)
    age = calc_age(bd, today) if bd else ""
    gender = (resp.gender if resp and resp.gender else (user.gender if user and user.gender else "")) or ""
    height = (getattr(resp, "height_cm", None) if resp else None) or (getattr(user, "height_cm", None) if user else None)
    weight = (getattr(resp, "weight_kg", None) if resp else None) or (getattr(user, "weight_kg", None) if user else None)

    # SurveyResponse 생성 (이번 제출 레코드에 NHIS를 '직접' 저장)
    sr = SurveyResponse(
        respondent_id=respondent_id,
        answers_json=json.dumps(normalized_payload, ensure_ascii=False),
        score=None,
        submitted_at=now_kst(),
        nhis_json=nhis_latest,   # 표준화된 작은 dict
        nhis_raw=nhis_raw,       # 원문 전체
    )
    session.add(sr)
    if resp:
        resp.status = "submitted"
        resp.updated_at = now_kst()
        session.add(resp)
    session.commit()
    session.refresh(sr)

    # 일련번호 채번
    if resp and resp.serial_no is None:
        next_val = session.exec(sa_text("SELECT nextval('respondent_serial_no_seq')")).one()[0]
        resp.updated_at = now_kst()
        resp.serial_no = next_val
        session.add(resp)
        session.commit()
        session.refresh(resp)

    # 건강검진 데이터 임시로그
    try:
        print("[NHIS][SAVE] latest_keys=", list((nhis_latest or {}).keys()) if isinstance(nhis_latest, dict) else type(nhis_latest))
    except Exception as e:
        print("[NHIS][SAVE][WARN1]", repr(e))
    ey = (nhis_latest.get("EXAMYEAR") or nhis_latest.get("GUNYEAR") or nhis_latest.get("exam_year"))
    print("[NHIS][SAVE] exam_year=", ey, "| has_latest:", bool(nhis_latest))

    # 자동 담당자 매핑 시도
    if resp:
        try_auto_map_partner_for_respondent(session, resp)

    # ── 알림메일: 매핑 완료된 경우에만 비동기 발송 ───────────────────────────────
    try:
        if resp and bool(getattr(resp, "is_mapped", False)):
            applicant_name = (
                (resp.applicant_name or (user.name_enc if user else ""))
                if resp else ""
            )
            created_at_kst_str = (
                to_kst(resp.created_at).strftime("%Y-%m-%d %H:%M")
                if resp and resp.created_at else now_kst().strftime("%Y-%m-%d %H:%M")
            )
            serial_no_val = resp.serial_no or 0

            background_tasks.add_task(
                send_submission_email,
                serial_no_val,
                applicant_name,
                created_at_kst_str,
            )
        else:
            # 매핑 전이면 메일 발송하지 않음 (매핑 완료 시점에 /partner/mapping에서 발송될 수 있음)
            print("[EMAIL] skip submission mail (not mapped yet): resp_id=", getattr(resp, "id", None))
    except Exception as e:
        print("[EMAIL][ERR]", repr(e))


    # 세션 정리 (작은 dict만 보관했었다면 이제 비워도 OK)
    request.session.pop("nhis_latest", None)
    request.session.pop("nhis_raw", None)

    #리다이렉트
    response = RedirectResponse(url="/portal", status_code=302)
    response.set_cookie(
        "survey_completed", "1",
        max_age=60*60*24*7,
        httponly=True,
        samesite="Lax",
        secure=bool(int(os.getenv("SECURE_COOKIE","1"))),
    )
    return response




@app.post("/admin/responses/export.xlsx")
async def admin_export_xlsx(
    request: Request,
    response: Response,
    ids: str = Form(...),
    session: Session = Depends(get_session),
    _auth: None = Depends(admin_required),
):
    # ---------------------------
    # 0) 유틸
    # ---------------------------
    def get_nhis_dict(v):
        """jsonb(dict) 또는 JSON 문자열 모두 수용"""
        if not v:
            return {}
        if isinstance(v, dict):
            return v
        try:
            return json.loads(v)
        except Exception:
            return {}

    def nhis_extract_all(nhis_json: dict | None, nhis_raw: dict | None) -> dict:
        """
        - 최우선: nhis_json(표준화된 최근 1건)에서 바로 꺼낸다.
        - 보조: nhis_raw.data.INCOMELIST가 있으면, 최신년도 1건으로 보강.
        - 결과: 엑셀 병합용 dict 리턴
        - 반환 키(영문):
          exam_year, exam_date, waist, osteoporosis,
          height, weight, bmi, bp, vision, hearing, hemoglobin, fbs, tc, hdl, ldl, tg,
          gfr, creatinine, ast, alt, ggt, urine_protein, chest, judgment
        """
        nj = nhis_json or {}

        def _year_of(d: dict) -> str:
            if not isinstance(d, dict):
                return ""
            for k in ("EXAMYEAR", "GUNYEAR", "YEAR", "YY"):
                v = d.get(k)
                if isinstance(v, int):
                    return str(v)
                if isinstance(v, str) and v.isdigit():
                    return v
            for k in ("EXAMDATE", "EXAM_DATE", "검진일자", "exam_date", "GUNDATE"):
                v = d.get(k)
                if isinstance(v, str) and len(v) >= 4 and v[:4].isdigit():
                    return v[:4]
            return ""

        def _mmdd_to_mm_dd(v) -> str:
            if v is None:
                return ""
            s = "".join(ch for ch in str(v) if ch.isdigit())
            if len(s) >= 4:
                return f"{s[:2]}-{s[2:4]}"
            return ""

        # 원본에서 최신 1건 추출 (INCOMELIST)
        raw_item = None
        if nhis_raw and isinstance(nhis_raw, dict):
            data = nhis_raw.get("data") or {}
            income = data.get("INCOMELIST") or []
            if isinstance(income, list) and income:
                def _yr(row: dict) -> int:
                    y = _year_of(row)
                    return int(y) if (y and y.isdigit()) else -1
                try:
                    income_sorted = sorted(income, key=_yr, reverse=True)
                except Exception:
                    income_sorted = list(income)
                raw_item = income_sorted[0] if income_sorted else None

        # 값 선택: 표준값 우선 → 없으면 원본 보강
        def pick(*keys: str) -> str:
            for k in keys:
                v = nj.get(k)
                if v not in (None, "", []):
                    return str(v)
            if raw_item:
                for k in keys:
                    v = raw_item.get(k)
                    if v not in (None, "", []):
                        return str(v)
            return ""

        # (요구사항) 검진년도: INCOMELIST.GUNYEAR 우선
        exam_year = ""
        if raw_item and raw_item.get("GUNYEAR") not in (None, "", []):
            exam_year = str(raw_item.get("GUNYEAR"))
        if not exam_year:
            exam_year = _year_of(nj) or _year_of(raw_item or {})

        # (요구사항) 검진일자: INCOMELIST.GUNDATE(mmdd) → "mm-dd"
        exam_date = ""
        if raw_item and raw_item.get("GUNDATE") not in (None, "", []):
            exam_date = _mmdd_to_mm_dd(raw_item.get("GUNDATE"))
        if not exam_date:
            exam_date = _mmdd_to_mm_dd(nj.get("GUNDATE") or nj.get("EXAMDATE") or nj.get("EXAM_DATE"))

        out = {
            "exam_year":      exam_year,
            "exam_date":      exam_date,
            "waist":          pick("WAISTSIZE"),
            "osteoporosis":   pick("OSTEOPOROSIS"),

            # 기존 항목들
            "height":         pick("HEIGHT"),
            "weight":         pick("WEIGHT"),
            "bmi":            pick("BODYMASS", "BMI"),
            "bp":             pick("BLOODPRESS"),
            "vision":         pick("SIGHT"),
            "hearing":        pick("HEARING"),
            "hemoglobin":     pick("HEMOGLOBIN"),
            "fbs":            pick("BLOODSUGAR"),   # 공복혈당
            "tc":             pick("TOTCHOLESTEROL"),
            "hdl":            pick("HDLCHOLESTEROL"),
            "ldl":            pick("LDLCHOLESTEROL"),
            "tg":             pick("TRIGLYCERIDE"),
            "gfr":            pick("GFR"),
            "creatinine":     pick("SERUMCREATININE"),
            "ast":            pick("SGOT"),
            "alt":            pick("SGPT"),
            "ggt":            pick("YGPT", "GAMMAGTP"),
            "urine_protein":  pick("YODANBAK"),
            "chest":          pick("CHESTTROUBLE"),
            "judgment":       pick("JUDGMENT"),
        }
        return out


    def fmt(v):
        if isinstance(v, list):
            return ",".join(str(x) for x in v)
        return "" if v is None else str(v)

    def calc_age(bd, ref_date):
        if not bd:
            return ""
        return ref_date.year - bd.year - ((ref_date.month, ref_date.day) < (bd.month, bd.day))

    # ---------------------------
    # 1) ids 파싱
    # ---------------------------
    print("export.xlsx ids raw:", repr(ids))
    id_list = [int(x) for x in (ids or "").split(",") if x.strip().isdigit()]
    if not id_list:
        return RedirectResponse(url="/admin/responses", status_code=303)

    # ---------------------------
    # 2) 질문 준비
    # ---------------------------
    def q_title(q: dict):
        return (q.get("title") or q.get("text") or q.get("label")
                or q.get("question") or q.get("prompt") or q.get("name")
                or f"Q{q.get('id','')}")
    questions_sorted = sorted(ALL_QUESTIONS, key=lambda x: x["id"])
    questions = [q_title(q) for q in questions_sorted]

    def extract_answers(payload: dict, questions_sorted: list[dict]) -> list:
        def to_num(v):
            if v is None or v == "": return ""
            if isinstance(v, list):
                return [int(x) for x in v if str(x).isdigit()]
            if isinstance(v, str) and "," in v:
                return [int(x) for x in v.split(",") if x.strip().isdigit()]
            return int(v) if str(v).isdigit() else ""

        ai = payload.get("answers_indices")
        if isinstance(ai, list) and len(ai) == len(questions_sorted):
            candidates = [payload]
            for k in ("answers", "acc", "data"):
                if isinstance(payload.get(k), dict):
                    candidates.append(payload[k])
            def fallback(qid):
                for cand in candidates:
                    v = cand.get(f"q{qid}")
                    num = to_num(v)
                    if num != "": return num
                return ""
            out = []
            for i, q in enumerate(questions_sorted):
                v = ai[i]
                empty = (v is None) or (v == "") or (isinstance(v, list) and len(v) == 0)
                out.append(fallback(q["id"]) if empty else v)
            return out

        candidates = [payload]
        for k in ("answers", "acc", "data"):
            if isinstance(payload.get(k), dict):
                candidates.append(payload[k])
        for cand in candidates:
            out, ok = [], 0
            for q in questions_sorted:
                num = to_num(cand.get(f"q{q['id']}"))
                out.append(num)
                if num != "": ok += 1
            if ok > 0:
                return out
        return ["" for _ in questions_sorted]

    # ---------------------------
    # 3) 워크북/시트 + 헤더
    # ---------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "문진결과"

    today = now_kst().date()

    base_headers = [
        "신청번호", "검진년도", "검진일자", "이름", "생년월일", "나이", "성별",
        "키", "체중", "BMI", "허리둘레", "혈압", "공복혈당", "총콜레스테롤", "중성지방",
        "HDL", "LDL", "혈색소(헤모글로빈)", "크레아티닌", "GFR", "AST", "ALT", "GGT",
    ]
    tail_headers = ["시력", "청력", "요단백", "흉부소견", "골밀도", "종합판정"]

    # 숫자 표시형식 적용 범위(키 ~ 문진 마지막)
    num_start_col = base_headers.index("키") + 1
    num_end_col = len(base_headers) + len(questions)

    ws.append(base_headers + questions + tail_headers)

    # ---------------------------
    # 4) 데이터 행
    # ---------------------------
    def to_number_cell(v):
        """가능하면 숫자 타입으로 변환 (정수는 int로 유지해서 17. 같은 표시 방지)"""
        if v is None or v == "":
            return None

        if isinstance(v, bool):
            return int(v)

        if isinstance(v, int):
            return v

        if isinstance(v, float):
            return int(v) if v.is_integer() else v

        if isinstance(v, list):
            if len(v) == 0:
                return None
            if len(v) == 1:
                return to_number_cell(v[0])
            # 다중선택은 숫자 강제 불가 -> 문자열
            parts = []
            for x in v:
                xx = to_number_cell(x)
                if xx is None:
                    continue
                parts.append(str(xx))
            return ",".join(parts)

        if isinstance(v, str):
            s = v.strip()

            # "17." 같은 형태 -> 정수로
            if s.endswith(".") and s[:-1].isdigit():
                return int(s[:-1])

            # "17.0", "17.00" -> 정수로
            if re.fullmatch(r"-?\d+(\.0+)", s):
                return int(s.split(".")[0])

            # 정수/소수
            if re.fullmatch(r"-?\d+", s):
                return int(s)

            if re.fullmatch(r"-?\d+\.\d+", s):
                f = float(s)
                return int(f) if f.is_integer() else f

        return v


    for idx, rid in enumerate(id_list, start=1):
        sr = session.get(SurveyResponse, rid)
        if not sr:
            continue

        resp = session.get(Respondent, sr.respondent_id) if sr.respondent_id else None
        user = session.get(User, resp.user_id) if resp and resp.user_id else None

        # 인적사항
        name = (resp.applicant_name if resp and resp.applicant_name else (user.name_enc if user and user.name_enc else "")) or ""
        bd = resp.birth_date if (resp and resp.birth_date) else (getattr(user, "birth_date", None) if user else None)
        age = calc_age(bd, today) if bd else ""
        gender = (resp.gender if resp and resp.gender else (user.gender if user and user.gender else "")) or ""
        serial_no = resp.serial_no if (resp and resp.serial_no is not None) else ""

        # 답 추출
        try:
            payload = json.loads(sr.answers_json) if sr.answers_json else {}
        except Exception as e:
            print("export.xlsx: bad answers_json for rid", rid, "err:", repr(e))
            payload = {}
        answers = extract_answers(payload, questions_sorted)

        # NHIS 표준+백업 추출 (표준: nhis_json, 원본: nhis_raw)
        nhis_std = nhis_extract_all(
            get_nhis_dict(sr.nhis_json),
            get_nhis_dict(sr.nhis_raw),
        )

        row = [
            serial_no,
            nhis_std.get("exam_year", ""),
            nhis_std.get("exam_date", ""),
            name,
            (bd.isoformat() if bd else ""),
            age,
            gender,

            # 키 ~ GGT (숫자 표시형식 대상)
            to_number_cell(nhis_std.get("height", "")),
            to_number_cell(nhis_std.get("weight", "")),
            to_number_cell(nhis_std.get("bmi", "")),
            to_number_cell(nhis_std.get("waist", "")),
            to_number_cell(nhis_std.get("bp", "")),
            to_number_cell(nhis_std.get("fbs", "")),
            to_number_cell(nhis_std.get("tc", "")),
            to_number_cell(nhis_std.get("tg", "")),
            to_number_cell(nhis_std.get("hdl", "")),
            to_number_cell(nhis_std.get("ldl", "")),
            to_number_cell(nhis_std.get("hemoglobin", "")),
            to_number_cell(nhis_std.get("creatinine", "")),
            to_number_cell(nhis_std.get("gfr", "")),
            to_number_cell(nhis_std.get("ast", "")),
            to_number_cell(nhis_std.get("alt", "")),
            to_number_cell(nhis_std.get("ggt", "")),
        ] + [to_number_cell(v) for v in answers] + [
            # 꼬리 열
            nhis_std.get("vision", ""),
            nhis_std.get("hearing", ""),
            nhis_std.get("urine_protein", ""),
            nhis_std.get("chest", ""),
            nhis_std.get("osteoporosis", ""),
            nhis_std.get("judgment", ""),
        ]

        ws.append(row)

    # (요구사항) "키" 열부터 "문진 마지막" 열까지 표시형식 = 숫자
    max_r = ws.max_row
    if max_r >= 2:
        for r in range(2, max_r + 1):
            for c in range(num_start_col, num_end_col + 1):
                cell = ws.cell(row=r, column=c)

                # 빈 값도 '숫자' 서식을 주되, 값 타입에 따라 표시형식 분기
                v = cell.value
                if isinstance(v, int):
                    cell.number_format = "0"
                elif isinstance(v, float):
                    if v.is_integer():
                        cell.value = int(v)
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.########"
                else:
                    # 숫자가 아닌 문자열(예: "120/80" 같은 혈압 문자열, 다중선택 "1,2")은 값 유지
                    # 그래도 요구사항상 범위는 숫자 열이므로, 빈칸(None)은 기본 숫자형 포맷을 줌
                    if v is None:
                        cell.number_format = "0"



    # ---------------------------
    # 5) 바이너리 응답
    # ---------------------------
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)
    return StreamingResponse(
        mem,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="responses.xlsx"'}
    )



@app.get("/nhis/progress")
def nhis_progress(request: Request, mode: str = "", callbackId: str = "", next: str = "/info"):
    # mode = 'callback' | 'direct'
    return templates.TemplateResponse(
        "nhis_progress.html",
        {
            "request": request,
            "mode": mode,
            "callbackId": callbackId,
            "next_url": next or "/info",
        },
    )



from fastapi import Body, Request, HTTPException

# ===========================================
# DataHub 간편인증 Step1: 시작
# ===========================================

@app.post("/api/dh/simple/start")
async def dh_simple_start(
    request: Request,
    session: Session = Depends(get_session),   # ★ 추가: 감사로그에 씁니다
):
    payload = await request.json()

    loginOption  = str(payload.get("loginOption", "")).strip()
    telecom      = str(payload.get("telecom", "")).strip()
    userName     = str(payload.get("userName", "")).strip()
    hpNumber     = str(payload.get("hpNumber", "")).strip()
    juminOrBirth = re.sub(r"[^0-9]", "", str(payload.get("juminOrBirth") or payload.get("birth") or ""))

    # 8자리 YYYYMMDD로 강제
    if len(juminOrBirth) >= 8:
        juminOrBirth = juminOrBirth[-8:]

    # ✅ LOGINOPTION 허용값: 0~7
    allowed = {"0","1","2","3","4","5","6","7"}

    #필수 입력값 점검
    missing = []
    if not loginOption or loginOption not in allowed:  missing.append("loginOption(0~7)")
    if not userName:                                   missing.append("userName")
    if not hpNumber:                                   missing.append("hpNumber")
    if not juminOrBirth:                               missing.append("birth(YYYYMMDD)")
    elif not re.fullmatch(r"\d{8}", juminOrBirth):     missing.append("birth(YYYYMMDD 8자리)") 
    if loginOption == "3" and not telecom:
        missing.append("telecom(PASS: 1~3, SKT|KT|LGU+)")

    if missing:
        logging.warning("[DH-START][VALIDATION] missing=%s", missing)
        return JSONResponse({"result":"FAIL","message":"필수 입력 누락","missing":missing}, status_code=400)

    # 새 트랜잭션 시작: 낡은 콜백/상태 제거  ← ★ 이 줄부터 추가
    for k in ("nhis_callback_id", "nhis_callback_type", "dh_callback"):
        request.session.pop(k, None)
    
    # hpNumber: 숫자만, 하이픈 없음
    hpNumber = re.sub(r'[^0-9]', '', hpNumber or '')

    # 콜백형 강제 규격 (LOGINOPTION 0~7 지원)
    dh_body = {
        "LOGINOPTION": loginOption,
        "HPNUMBER":    hpNumber,
        "USERNAME":    userName,
        "JUMIN":       juminOrBirth,
    }
    if loginOption == "3" and telecom:
        dh_body["TELECOMGUBUN"] = telecom  # 1~3
    
    # (선택) 민감값 마스킹 로그
    _safe = {**dh_body, "HPNUMBER": _mask_phone(dh_body.get("HPNUMBER","")), "JUMIN": _mask_birth(dh_body.get("JUMIN",""))}
    logging.debug("[DH-START][BODY]%s", _safe)
    
    #성별 세션 보관
    gender = str(payload.get("gender","")).strip() 
    request.session["nhis_gender"] = gender if gender in ("남","여") else ""
    
    # ── 동의 여부 파싱 ──────────────────────────────────────────────
    agreement_all      = bool(payload.get("agreementAll"))
    agreement_collect  = bool(payload.get("agreementCollect"))
    agreement_third    = bool(payload.get("agreementThird"))
    agreement_unique   = bool(payload.get("agreementUnique"))
    agreement_overseas = bool(payload.get("agreementOverseas"))

    # 세션에 동의 여부 + 동의 시각만 기록 (DB 반영은 /survey/finish 에서)
    request.session["agreement_all"] = agreement_all
    if agreement_all:
        # 세션은 datetime을 그대로 못 넣으니까 문자열로 저장
        request.session["agreement_at"] = now_kst().isoformat()
    else:
        # 전체 동의 취소한 경우 흔적 제거
        request.session.pop("agreement_at", None)

    # 인적정보 세션 보관 (원래 있던 줄은 그대로 유지)
    request.session["nhis_start_payload"] = dh_body

        

    # ===============================================
    # 1) DataHub.simple_auth_start 재시도(최대 3회)
    # ===============================================
    rsp = None
    last_error = None

    for attempt in range(1, 4):
        try:
            logging.info(
                "[DH][START][TRY] attempt=%s LOGINOPTION=%s name=%s phone=%s",
                attempt,
                dh_body.get("LOGINOPTION"),
                dh_body.get("USERNAME"),
                dh_body.get("HPNUMBER"),
            )

            rsp = DATAHUB.simple_auth_start(
                login_option=dh_body["LOGINOPTION"],
                user_name=dh_body["USERNAME"],
                hp_number=dh_body["HPNUMBER"],
                jumin_or_birth=dh_body["JUMIN"],
                telecom_gubun=dh_body.get("TELECOMGUBUN"),
            )

            # 정상 응답 → 재시도 중단
            break

        except DatahubError as e:
            last_error = e
            logging.warning(
                "[DH][START][RETRY] attempt=%s error=%r",
                attempt, e
            )

            if attempt >= 3:
                msg = "현재 국가건강검진 조회 서비스 연결이 불안정하여 연결에 실패하였습니다.\n잠시 후 다시 시도해주세요."
                logging.error(
                    "[DH][START][TIMEOUT] attempts=3 last_error=%r",
                    last_error,
                )
                return JSONResponse(
                    status_code=200,
                    content={
                        "errCode": "NETWORK_TIMEOUT",
                        "message": msg,
                        "data": None
                    }
                )
            time.sleep(0.3)


    try:
        stmt = sa_text("""
            INSERT INTO nhis_audit (respondent_id, callback_id, request_json, response_json)
            VALUES (:rid, :cbid, :req, :res)
        """).bindparams(
            rid=None,  # 알 수 없으면 None
            cbid=(rsp.get("data") or {}).get("callbackId") or "",
            req=json.dumps({"step":"start","body":dh_body}, ensure_ascii=False),
            res=json.dumps(rsp or {}, ensure_ascii=False),
        )
        session.exec(stmt)
        session.commit()
    except Exception as e:
        print("[NHIS][AUDIT][ERR][start]", repr(e))


    err  = str(rsp.get("errCode","")).strip()
    data = rsp.get("data") or {}
    cbid = (data.get("callbackId") or "").strip()

    if err == "0001" and cbid:
        request.session["nhis_callback_id"]   = cbid
        request.session["nhis_callback_type"] = "SIMPLE"
        return JSONResponse({"errCode":"0001","message":"NEED_CALLBACK","data":{"callbackId": cbid}}, status_code=200)

    # ★ 여기서부터는 전부 실패로 처리 (0000이라도 콜백 없으면 실패)
    msg = (rsp.get("errMsg") or "간편인증 시작 실패").strip()
    return JSONResponse({"errCode": err or "9999", "message": msg, "data": data}, status_code=200)



# ===========================================
# DataHub 간편인증 Step2: 완료(captcha)
# ===========================================


@app.post("/api/dh/simple/complete")
async def dh_simple_complete(
    request: Request,
    session: Session = Depends(get_session),
):
    """
    콜백형 완료:
      1) /scrap/captcha (Step2) 1회 호출 (callbackResponse* 키 포함, 예외/타임아웃 안전 처리)
      2) Step2 응답에 INCOMELIST가 있으면 즉시 채택하여 종료
      3) 없으면 같은 callbackId로 /scrap/common/...Simple 폴링 재조회 (light만)
      4) 최대 120초 폴링, 미완료면 202
    """

    try:
        # JSON body가 있으면 파싱, 없으면 그냥 빈 dict
        payload = await request.json()
        if not isinstance(payload, dict):
            payload = {}
    except Exception:
        payload = {}

    # 0) 세션(or 요청)에서 콜백 값 복구
    cbid = (request.session or {}).get("nhis_callback_id") or str(payload.get("callbackId") or "")
    cbtp = (request.session or {}).get("nhis_callback_type") or str(payload.get("callbackType") or "SIMPLE")

    # callbackid/type 확인 디버그 로그
    logging.debug(
        "[DH-COMPLETE][DEBUG] callbackId sources => session:%s | payload:%s | final:%s",
        (request.session or {}).get("nhis_callback_id"),
        payload.get("callbackId"),
        cbid,
    )
    logging.debug(
        "[DH-COMPLETE][DEBUG] callbackType sources => session:%s | payload:%s | final:%s",
        (request.session or {}).get("nhis_callback_type"),
        payload.get("callbackType"),
        cbtp,
    )

    # 해당 시점 rtoken 유무 확인 디버그 로그
    try:
        rtok = (request.query_params.get("rtoken") or request.cookies.get("rtoken") or "")
        rid_dbg = verify_token(rtok) if rtok else None
        logging.debug("[RTOKEN][DBG][complete] raw=%s | rid=%s", ("yes" if rtok else "no"), (rid_dbg if rid_dbg else "None"))
    except Exception as e:
        logging.debug("[RTOKEN][DBG][complete][ERR] %r", e)


    # 0-1) 최소 검증 (DataHub 호출 낭비 방지)
    if not cbid or not cbtp:
        logging.warning("[DH-COMPLETE][VALIDATION] missing=%s", [k for k, v in {"callbackId": cbid, "callbackType": cbtp}.items() if not v])
        return JSONResponse({"result": "FAIL", "message": "필수 입력 누락", "missing": ["callbackId", "callbackType"]}, status_code=400)

    # 1) Step2: /scrap/captcha (키는 모두 포함; 값은 비어도 OK) — 예외/타임아웃 안전 처리
    try:
        step2_res = DATAHUB.simple_auth_complete(
            callback_id=cbid,
            callback_type=cbtp,
            callbackResponse=str(payload.get("callbackResponse") or ""),
            callbackResponse1=str(payload.get("callbackResponse1") or ""),
            callbackResponse2=str(payload.get("callbackResponse2") or ""),
            retry=str(payload.get("retry") or ""),
        )
    except Exception as e:
        # 네트워크/타임아웃 등 오류가 나더라도 폴링으로 진행
        print("[DH-COMPLETE][captcha][ERR]", repr(e))
        step2_res = {"errCode": "TIMEOUT", "result": "FAIL", "data": {}}

    # 1-1) 감사로그: Step2 요청/응답 저장
    try:
        resp_id = None
        try:
            tok = (request.query_params.get("rtoken") or request.cookies.get("rtoken") or "")
            rid = verify_token(tok) if tok else -1
            if rid > 0:
                resp_id = rid
        except Exception:
            pass

        stmt = sa_text("""
            INSERT INTO nhis_audit (respondent_id, callback_id, request_json, response_json)
            VALUES (:rid, :cbid, :req, :res)
        """).bindparams(
            rid=resp_id,
            cbid=cbid,
            req=json.dumps({"step": "captcha", "body": {
                "callbackId": cbid, "callbackType": cbtp,
                "callbackResponse": "", "callbackResponse1": "", "callbackResponse2": "", "retry": ""
            }}, ensure_ascii=False),
            res=json.dumps(step2_res or {}, ensure_ascii=False),
        )
        session.exec(stmt)
        session.commit()
    except Exception as e:
        print("[NHIS][AUDIT][ERR][captcha]", repr(e))

    # 1-2) Step2 응답 자체에 INCOMELIST가 있으면 즉시 채택
    try:
        step2_data = (step2_res or {}).get("data") or {}
        income2 = step2_data.get("INCOMELIST") or []
        if isinstance(income2, list) and len(income2) > 0:
            want_all = (request.query_params.get("all") or "").lower() in ("1", "true", "yes")
            picked = pick_latest_general(step2_res, mode=("all" if want_all else "latest"))
            request.session["nhis_latest"] = picked if isinstance(picked, dict) else {}

            # ★ NHIS결과 DB 저장 (엑셀 병합용)
            try:
                picked_one = pick_latest_general(step2_res, mode="latest")
                _save_nhis_to_db(session, request, picked_one, step2_res)
                # 저장이 스킵될 수도 있으니 세션에도 임시 보관
                request.session["nhis_latest"] = picked_one or {}
                #request.session["nhis_raw"]    = step2_res or {}  #쿠키 터짐 저장하지 않음
            except Exception as e:
                print("[NHIS][DB][WARN][captcha-save]", repr(e))


            # --- 성공 직전 User 인적정보 업데이트(이름/성별/생년월일) ---
            try:
                auth_cookie = request.cookies.get(AUTH_COOKIE_NAME)
                user_id = verify_user(auth_cookie) if auth_cookie else -1
                if user_id and user_id > 0:
                    user = session.get(User, user_id)
                    if user:
                        sp = (request.session or {}).get("nhis_start_payload") or {}
                        nm = str(sp.get("USERNAME") or "").strip()
                        bd8 = str(sp.get("JUMIN") or "").strip()
                        gd = (request.session or {}).get("nhis_gender") or ""
                        # 생년월일 파싱(YYYYMMDD)
                        bd_date = None
                        if len(bd8) == 8 and bd8.isdigit():
                            bd_date = date(int(bd8[0:4]), int(bd8[4:6]), int(bd8[6:8]))
                        # 저장(있을 때만 덮어씀)
                        if nm: user.name_enc = nm
                        if gd in ("남","여"): user.gender = gd
                        if bd_date: user.birth_date = bd_date; user.birth_year = bd_date.year
                        session.add(user); session.commit()
            except Exception as _e:
                logging.debug("[NHIS][USER-SNAPSHOT][WARN] %r", _e)

            return JSONResponse({"ok": True, "errCode": "0000", "message": "OK", "data": picked}, status_code=200)
  
    except Exception as e:
        print("[DH-COMPLETE][WARN][captcha-pick]", repr(e))

    # 2) 결과 재조회 폴링 (light 1회 확인 후 → full만)
    max_wait_sec = NHIS_POLL_MAX_SEC
    deadline = time.time() + max_wait_sec
    attempt = 0

    # 시작 단계 값 복구
    SP = (request.session or {}).get("nhis_start_payload") or {}
    loginOption  = str(SP.get("LOGINOPTION", "")).strip()
    userName     = str(SP.get("USERNAME", "")).strip()
    hpNumber     = str(SP.get("HPNUMBER", "")).strip()
    juminVal     = str(SP.get("JUMIN", "")).strip() or str(SP.get("JUMINNUM", "")).strip()
    telecomGubun = str(SP.get("TELECOMGUBUN", "")).strip() if loginOption == "3" else None

    want_all = ((request.query_params.get("all") or "").lower() in ("1", "true", "yes"))

    did_full = False
    empty_income_streak = 0
    tried_identity = 0
    while time.time() < deadline:
        attempt += 1
        try:
            # ✅ 가이드에 맞게: Step2 이후 재조회도 callback 기반(light)만 사용
            #    (full=신상정보 포함 재요청은 Step1 성격이라 0001 재발 가능)
            fetch_body = {"CALLBACKID": cbid, "CALLBACKTYPE": cbtp}
            rsp2 = DATAHUB.medical_checkup_simple(fetch_body)
            kind = "light"
        except Exception as e:
            logging.warning("[DH-COMPLETE][FETCH][ERR] %r", e)
            time.sleep(NHIS_FETCH_INTERVAL)
            continue

        # 감사로그(요약)
        try:
            stmt = sa_text("""
                INSERT INTO nhis_audit (respondent_id, callback_id, request_json, response_json)
                VALUES (:rid, :cbid, :req, :res)
            """).bindparams(
                rid=None, cbid=cbid,
                req=json.dumps({"step": "fetch", "kind": kind}, ensure_ascii=False),
                res=json.dumps(rsp2 or {}, ensure_ascii=False),
            )
            session.exec(stmt); session.commit()
        except Exception as e:
            logging.warning("[NHIS][AUDIT][ERR][fetch-log] %r", e)

        err2   = str((rsp2 or {}).get("errCode") or "")
        data2  = (rsp2 or {}).get("data") or {}
        income = data2.get("INCOMELIST") or []

        # ✅ 10년 중 데이터 없음 케이스: errCode=0000 success인데 INCOMELIST가 끝까지 비는 경우
        # 가이드 예시처럼 data.RESULT가 FAIL로 내려오는 패턴을 "데이터 없음"으로 간주
        try:
            result_flag = str(data2.get("RESULT") or "").upper()
        except Exception:
            result_flag = ""

        if (
            err2 == "0000"
            and isinstance(income, list)
            and len(income) == 0
            and result_flag == "FAIL"
        ):
            msg = "과거 10년 중 국가건강검진 데이터가 존재하지 않습니다. 담당자에게 문의해주세요"
            return JSONResponse(
                status_code=200,
                content={
                    "ok": False,
                    "errCode": "NO_DATA",
                    "msg": msg,
                    "message": msg,
                    "data": data2,
                },
            )


        # 내부 에러 힌트만 DEBUG로
        inner_ecode  = data2.get("ECODE")
        inner_errmsg = data2.get("ERRMSG")
        if inner_ecode and inner_ecode != "0000":
            logging.debug("[DH-COMPLETE][FETCH][INNER] ecode=%s msg=%s", inner_ecode, inner_errmsg)

        logging.info("[DH-COMPLETE][FETCH] attempt=%s kind=%s err=%s income_len=%s",
                     attempt, kind, err2, (len(income) if isinstance(income, list) else "NA"))

        # ★ errCode 2003: 이용횟수 소진 → 재시도해도 소용 없으므로 즉시 종료
        if err2 == "2003":
            msg = (
                data2.get("ERRMSG")
                or (rsp2 or {}).get("errMsg")
                or "건강검진 조회 가능 횟수가 소진되었습니다. 관리자에게 문의해주세요."
            )
            return JSONResponse(
                status_code=200,
                content={
                    "ok": False,
                    "errCode": err2,
                    "msg": msg,
                    "message": msg,
                    "data": data2,
                },
            )

        # ✅ callback 조회는 성공(0000)인데 INCOMELIST가 계속 비는 경우가 있음
        #    이때만 identity 기반 조회를 "최대 2회" 제한적으로 fallback 시도
        if err2 == "0000" and isinstance(income, list) and len(income) == 0:
            empty_income_streak += 1
        else:
            empty_income_streak = 0

        if (
            err2 == "0000"
            and isinstance(income, list)
            and len(income) == 0
            and empty_income_streak >= 5
            and tried_identity < 2
        ):
            tried_identity += 1
            logging.info("[DH-COMPLETE][FALLBACK] try identity fetch #%s (empty_income_streak=%s)", tried_identity, empty_income_streak)

            # 시작 단계 값 복구(이미 위에서 SP/loginOption 등 복구하는 코드가 있으면 그걸 그대로 사용)
            SP = (request.session or {}).get("nhis_start_payload") or {}
            loginOption  = str(SP.get("LOGINOPTION", "")).strip()
            userName     = str(SP.get("USERNAME", "")).strip()
            hpNumber     = str(SP.get("HPNUMBER", "")).strip()
            juminVal     = str(SP.get("JUMIN", "")).strip() or str(SP.get("JUMINNUM", "")).strip()
            telecomGubun = str(SP.get("TELECOMGUBUN", "")).strip() if loginOption == "3" else None

            try:
                rsp_id = DATAHUB.medical_checkup_simple_with_identity(
                    callback_id=cbid,
                    callback_type=cbtp,
                    login_option=loginOption,
                    user_name=userName,
                    hp_number=hpNumber,
                    jumin_or_birth=juminVal,
                    telecom_gubun=telecomGubun
                )
                err_id = str((rsp_id or {}).get("errCode") or "")
                data_id = (rsp_id or {}).get("data") or {}
                income_id = data_id.get("INCOMELIST") or []

                logging.info("[DH-COMPLETE][FALLBACK] identity err=%s income_len=%s", err_id, (len(income_id) if isinstance(income_id, list) else "NA"))

                # identity에서 0000 + income 있으면 그걸로 성공 처리
                if err_id == "0000" and isinstance(income_id, list) and len(income_id) > 0:
                    picked = pick_latest_general(rsp_id, mode=("all" if want_all else "latest"))
                    request.session["nhis_latest"] = picked if isinstance(picked, dict) else {}
                    try:
                        picked_one = pick_latest_general(rsp_id, mode="latest")
                        _save_nhis_to_db(session, request, picked_one, rsp_id)
                        request.session["nhis_latest"] = picked_one or {}
                    except Exception as e:
                        logging.warning("[NHIS][DB][WARN][identity-save] %r", e)

                    return JSONResponse({"ok": True, "errCode": "0000", "message": "OK", "data": picked}, status_code=200)

                # identity에서 0001이면(추가 텍스트 요구) → fallback 중단하고 callback 폴링 계속
                if err_id == "0001":
                    logging.info("[DH-COMPLETE][FALLBACK] identity returned 0001; continue callback polling")

            except Exception as e:
                logging.warning("[DH-COMPLETE][FALLBACK][ERR] %r", e)

            # fallback 이후에는 바로 다음 폴링 주기로
            time.sleep(NHIS_FETCH_INTERVAL)
            continue



        if err2 == "0000" and isinstance(income, list) and len(income) > 0:
            picked = pick_latest_general(rsp2, mode=("all" if want_all else "latest"))
            request.session["nhis_latest"] = picked if isinstance(picked, dict) else {}
            # DB 저장 (엑셀 병합용)
            try:
                picked_one = pick_latest_general(rsp2, mode="latest")
                _save_nhis_to_db(session, request, picked_one, rsp2)
                request.session["nhis_latest"] = picked_one or {}
            except Exception as e:
                logging.warning("[NHIS][DB][WARN][fetch-save] %r", e)
                
            # --- 성공 직전 User 인적정보 업데이트(이름/성별/생년월일) ---
            try:
                auth_cookie = request.cookies.get(AUTH_COOKIE_NAME)
                user_id = verify_user(auth_cookie) if auth_cookie else -1
                if user_id and user_id > 0:
                    user = session.get(User, user_id)
                    if user:
                        sp = (request.session or {}).get("nhis_start_payload") or {}
                        nm = str(sp.get("USERNAME") or "").strip()
                        bd8 = str(sp.get("JUMIN") or "").strip()
                        gd = (request.session or {}).get("nhis_gender") or ""
                        # 생년월일 파싱(YYYYMMDD)
                        bd_date = None
                        if len(bd8) == 8 and bd8.isdigit():
                            bd_date = date(int(bd8[0:4]), int(bd8[4:6]), int(bd8[6:8]))
                        # 저장(있을 때만 덮어씀)
                        if nm: user.name_enc = nm
                        if gd in ("남","여"): user.gender = gd
                        if bd_date: user.birth_date = bd_date; user.birth_year = bd_date.year
                        session.add(user); session.commit()
            except Exception as _e:
                logging.debug("[NHIS][USER-SNAPSHOT][WARN] %r", _e)
            
            return JSONResponse({"ok": True, "errCode": "0000", "message": "OK", "data": picked}, status_code=200)

        time.sleep(NHIS_FETCH_INTERVAL)


    return JSONResponse({"ok": False, "errCode": "2020", "message": "아직 인증이 완료되지 않았거나 데이터가 준비되지 않았습니다."}, status_code=202)




# ---- 유틸: 최신 1건 선택 (연/월/일 기준) ----
def pick_latest_one(data: dict) -> dict:
    """
    data.INCOMELIST[] 중 가장 최근(연/월/일) 1건만 골라 요약해 리턴.
    형식은 가이드의 필드명을 그대로 사용.
    """
    items = (data or {}).get("502INCOMELIST") or []
    best = None
    best_key = None
    for it in items:
        year = (it.get("GUNYEAR") or "").strip()
        date = (it.get("GUNDATE") or "").strip()  # 'MM/DD' 형태 예시
        # 키를 YYYYMMDD 정수로 만들어 비교
        try:
            mm, dd = (date.split("/") + ["0", "0"])[:2]
            key = int(f"{int(year):04d}{int(mm):02d}{int(dd):02d}")
        except Exception:
            # 연도만 있는 항목은 월/일 0으로
            try:
                key = int(f"{int(year):04d}0000")
            except Exception:
                key = -1
        if best is None or key > best_key:
            best, best_key = it, key
    return best or {}



# ===========================================
# DataHub 인증서 방식(필요 시): 건강검진 결과 조회
# ===========================================
@app.post("/api/dh/nhis/result")
def dh_nhis_result(payload: dict = Body(...), request: Request = None):
    """
    요청 JSON:
    {
      "jumin": "주민번호13자리",
      "certName": "cn=...,ou=...,o=...,c=kr",
      "certPwd": "인증서비번",
      "derB64": "<DER base64>",
      "keyB64": "<KEY base64>"
    }
    """
    try:
        jumin   = (payload.get("jumin") or "").strip()
        cname   = (payload.get("certName") or "").strip()
        cpwd    = (payload.get("certPwd") or "").strip()
        der_b64 = (payload.get("derB64") or "").strip()
        key_b64 = (payload.get("keyB64") or "").strip()
        if not (jumin and cname and cpwd and der_b64 and key_b64):
            raise HTTPException(400, "jumin/certName/certPwd/derB64/keyB64는 모두 필수입니다.")
        rsp = DATAHUB.nhis_medical_checkup(jumin, cname, cpwd, der_b64, key_b64)
        try:
            picked = pick_latest_general(rsp)
            if request is not None:
                request.session["nhis_latest"] = picked
        except Exception:
            pass
        return rsp
    except DatahubError as e:
        raise HTTPException(502, f"DATAHUB error: {e}")
    except Exception as e:
        raise HTTPException(500, f"Internal error: {e}")


#임시 디버그 라우트, 로그. 운영 시 삭제
@app.get("/debug/datahub-selftest")
def debug_datahub_selftest():
    from app.vendors.datahub_client import encrypt_field, _get_key_iv, _get_text_encoding

    plain  = (os.getenv("DATAHUB_SELFTEST_PLAIN", "") or "").strip()
    expect = (os.getenv("DATAHUB_SELFTEST_EXPECT", "") or "").strip()
    if not plain or not expect:
        return JSONResponse({"error":"set DATAHUB_SELFTEST_PLAIN & EXPECT"}, status_code=400)

    # 현재 FORCE 설정/키/IV 요약
    kb  = int(os.getenv("DATAHUB_FORCE_KEY_BITS", "256") or "256")
    ivm = (os.getenv("DATAHUB_FORCE_IV_MODE", "ENV") or "ENV").upper()
    ksh = (os.getenv("DATAHUB_FORCE_KEY_SHAPE", "right") or "right").lower()
    enc = _get_text_encoding()

    key, iv = _get_key_iv()
    key_sha = hashlib.sha256(key).hexdigest()[:16]
    iv_hex  = iv.hex()[:16]

    got = encrypt_field(plain)
    return JSONResponse({
        "env": {
            "DATAHUB_TEXT_ENCODING": enc,
            "DATAHUB_FORCE_KEY_BITS": kb,
            "DATAHUB_FORCE_IV_MODE": ivm,
            "DATAHUB_FORCE_KEY_SHAPE": ksh
        },
        "kiv": {
            "key_bits": kb,
            "key_len": len(key),
            "key_sha256_head": key_sha,
            "iv_len": len(iv),
            "iv_head_hex": iv_hex
        },
        "plain": plain,
        "expect": expect,
        "got": got,
        "match": (got == expect)
    })



#임시 디버그 라우트, 로그. 운영 시 삭제
@app.get("/debug/whoami")
def debug_whoami():
    return JSONResponse({
        "main_file": __file__,
        "cwd": os.getcwd(),
        "env": {
            "APP_ENV": os.getenv("APP_ENV"),
            "DATAHUB_SELFTEST": os.getenv("DATAHUB_SELFTEST"),
            "DATAHUB_SELFTEST_PLAIN_set": bool(os.getenv("DATAHUB_SELFTEST_PLAIN")),
            "DATAHUB_SELFTEST_EXPECT_set": bool(os.getenv("DATAHUB_SELFTEST_EXPECT")),
        }
    })

#인코딩 테스트
@app.get("/debug/datahub-finder")
def debug_datahub_finder():
    """
    API 호출 없이, ENV에 있는 PlainData/EncData 쌍을 기준으로
    - 평문 인코딩(utf-8/cp949)
    - 키 비트(128/256)
    - IV 모드(ENV/ZERO)
    - EncKey/IV 해석(., urlsafe, pad, hex, raw 등 변형)
    - 키/IV 길이 보정 방식(left/right pad)
    - (보너스) 키 유도(SHA-256/MD5) 방식
    조합을 브루트포스로 시도해 'expect'와 일치하는 암호문을 찾는다.
    """
    def pkcs7_pad(b: bytes, block=16) -> bytes:
        n = block - (len(b) % block)
        return b + bytes([n]) * n

    plain  = (os.getenv("DATAHUB_SELFTEST_PLAIN", "") or "").strip()
    expect = (os.getenv("DATAHUB_SELFTEST_EXPECT", "") or "").strip()
    enc_spec = (os.getenv("DATAHUB_ENC_SPEC", "") or "").strip().upper()
    enc_key_raw = (os.getenv("DATAHUB_ENC_KEY_B64", "") or "").strip()
    enc_iv_raw  = (os.getenv("DATAHUB_ENC_IV_B64", "") or "").strip()

    if not plain or not expect or not enc_key_raw:
        return JSONResponse({"error":"set DATAHUB_SELFTEST_PLAIN/EXPECT and ENC_KEY/IV"}, status_code=400)

    # 1) key/iv 해석 후보 생성
    def pad4(s: str) -> str:
        return s + ("=" * ((4 - len(s) % 4) % 4))

    def b64_try(s: str):
        cands = [
            s,
            s.replace("-", "+").replace("_", "/"),
            s.replace(".", ""),           # dot 제거
            s.replace(".", "+"),
            s.replace(".", "/"),
        ]
        seen = set()
        out = []
        for c in cands:
            c = pad4(c)
            if c in seen: continue
            seen.add(c)
            try:
                out.append(base64.b64decode(c))
            except Exception:
                pass
        return out

    def hex_try(s: str):
        try:
            return [bytes.fromhex(s)]
        except Exception:
            return []

    def raw_try(s: str):
        return [s.encode("utf-8")]

    key_bytes_cands = b64_try(enc_key_raw) + hex_try(enc_key_raw) + raw_try(enc_key_raw)
    iv_bytes_cands  = b64_try(enc_iv_raw)  + hex_try(enc_iv_raw)  + raw_try(enc_iv_raw)
    if not iv_bytes_cands:
        iv_bytes_cands = [b"\x00"*16]  # IV 미제공 대비

    # 2) 길이 보정/유도 함수들
    def shape_key(k: bytes, bits: int, mode: str) -> bytes:
        need = 32 if bits == 256 else 16
        if mode == "right":
            return (k[:need]).ljust(need, b"\x00")
        elif mode == "left":
            return (k[-need:]).rjust(need, b"\x00")
        elif mode == "sha256":
            d = hashlib.sha256(k).digest()
            return d[:need]
        elif mode == "md5":
            d = hashlib.md5(k).digest()
            # 128비트만 직접 충족, 256은 md5 두 번 접합
            return (d if need==16 else (d+hashlib.md5(d).digest()))[:need]
        else:
            return (k[:need]).ljust(need, b"\x00")

    def shape_iv(iv: bytes, mode: str) -> bytes:
        if mode == "ZERO":
            return b"\x00"*16
        v = iv[:16]
        if len(v) < 16: v = v.ljust(16, b"\x00")
        return v

    # 3) 시도할 조합들
    encodings   = ["utf-8", "cp949"]
    key_bits    = [256, 128]
    key_shapes  = ["right", "left", "sha256", "md5"]  # 키 길이/유도 방식
    iv_modes    = ["ENV", "ZERO"]
    iv_shapes   = ["keep"]  # 필요 시 확장
    tried = 0
    matches = []

    for enc in encodings:
        p = plain.encode(enc, errors="strict")
        for kb in key_bits:
            for ks in key_shapes:
                for key0 in key_bytes_cands:
                    key = shape_key(key0, kb, ks)
                    for ivm in iv_modes:
                        for ivs in iv_shapes:
                            for iv0 in iv_bytes_cands:
                                iv = shape_iv(iv0, ivm)
                                data = pkcs7_pad(p, 16)
                                try:
                                    ct = AES.new(key, AES.MODE_CBC, iv).encrypt(data)
                                    b64 = base64.b64encode(ct).decode("ascii")
                                except Exception:
                                    continue
                                tried += 1
                                if b64 == expect:
                                    matches.append({
                                        "encoding": enc,
                                        "key_bits": kb,
                                        "key_shape": ks,
                                        "iv_mode": ivm,
                                        "iv_note": ("ENV" if ivm=="ENV" else "ZERO"),
                                        "key_src_len": len(key0),
                                        "iv_src_len": len(iv0),
                                        "b64": b64
                                    })
                                    # 바로 리턴(첫 일치)
                                    return JSONResponse({
                                        "status":"MATCH",
                                        "tried": tried,
                                        "enc_spec": enc_spec,
                                        "match": matches[0]
                                    })

    return JSONResponse({"status":"NO_MATCH", "tried": tried, "enc_spec": enc_spec})


def ensure_nhis_audit_table():
    try:
        with Session(engine) as s:
            s.exec(text("""
            CREATE TABLE IF NOT EXISTS nhis_audit (
              id BIGSERIAL PRIMARY KEY,
              created_at timestamptz DEFAULT now(),
              stage text,             -- start / complete / fetch
              callback_id text,
              err_code text,
              user_mask text,
              rsp_json jsonb
            );
            """))
            s.commit()
    except Exception as e:
        print("[NHIS][AUDIT][WARN]", repr(e))

def audit_nhis(stage, err_code, callback_id, rsp_json=None, user_mask=None):
    try:
        with Session(engine) as s:
            s.exec(text("""
            INSERT INTO nhis_audit (stage, err_code, callback_id, user_mask, rsp_json)
            VALUES (:stage, :err, :cbid, :mask, :rsp)
            """), {"stage":stage, "err":err_code, "cbid":callback_id, "mask":user_mask or "", "rsp":json.dumps(rsp_json or {})})
            s.commit()
    except Exception as e:
        print("[NHIS][AUDIT][ERR]", repr(e))


@app.get("/_routes")
def _routes():
    return [{"path": r.path, "methods": list(r.methods)} for r in app.routes if isinstance(r, APIRoute)]


#슬러그 기반 랜딩 라우트 추가: 파일 가장 아래(라우트들의 가장 마지막)에 두는게 안전.
#그래야 /login, /partner/login 같은 고정 라우트가 먼저 매칭돼서 충돌이 안 남)
# =========================================================
# Partner slug landing routes (MUST be the LAST routes)
# =========================================================

@app.get("/{p_slug}", response_class=HTMLResponse)
def partner_landing(
    request: Request,
    p_slug: str,
    session: Session = Depends(get_session),
):
    # slug 형식 제한 (운영 안전)
    if not re.fullmatch(r"[a-z0-9-]{2,80}", p_slug or ""):
        raise HTTPException(status_code=404)

    p = get_partner_by_slug(session, p_slug)  # 내부에서 is_active=TRUE 조건 적용
    if not p or not p.get("p_name"):
        raise HTTPException(status_code=404)

    # index.html 재사용 (외형 동일) - 버튼만 start로 연결
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "apply_url": f"/{p_slug}/start"},
    )


@app.get("/{p_slug}/start")
def partner_landing_start(
    request: Request,
    p_slug: str,
    session: Session = Depends(get_session),
):
    if not re.fullmatch(r"[a-z0-9-]{2,80}", p_slug or ""):
        raise HTTPException(status_code=404)

    p = get_partner_by_slug(session, p_slug)  # 내부에서 is_active=TRUE 조건 적용
    if not p or not p.get("p_name"):
        raise HTTPException(status_code=404)

    CO_CAMPAIGN_ID = p["p_name"]  # campaign_id는 partner.p_name

    # 기존 코드/담당자 유입 세션 흔적 제거
    request.session.pop("partner_id", None)
    request.session.pop("admin_phone", None)

    # 캠페인 세션 저장 (survey_root에서 respondent.campaign_id로 저장)
    request.session["campaign_id"] = CO_CAMPAIGN_ID

    # AUTH 쿠키 발급용 임시 유저 생성
    ph = f"{p_slug}_{secrets.token_hex(16)}"
    user = User(phone_hash=ph)
    session.add(user)
    session.commit()
    session.refresh(user)

    SECURE_COOKIE = bool(int(os.getenv("SECURE_COOKIE", "1")))
    resp = RedirectResponse(url="/nhis", status_code=303)
    resp.set_cookie(
        AUTH_COOKIE_NAME,
        sign_user(user.id),
        httponly=True,
        secure=SECURE_COOKIE,
        samesite="lax",
        max_age=AUTH_MAX_AGE,
    )
    resp.delete_cookie("survey_completed")

    # respondent 미리 생성 + partner_id는 매핑 전이므로 None
    r = Respondent(
        user_id=user.id,
        campaign_id=CO_CAMPAIGN_ID,
        status="started",
        partner_id=None,
    )
    session.add(r)
    session.commit()
    session.refresh(r)

    request.session["respondent_id"] = r.id
    return resp
